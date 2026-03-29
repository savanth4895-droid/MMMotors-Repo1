from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
import asyncio
import aiofiles
import zipfile
import shutil
import json
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta, timezone
import bcrypt
import jwt
from enum import Enum
from fastapi import UploadFile, File
import csv
import io
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection with Atlas-compatible settings
mongo_url = os.environ['MONGO_URL']

# Detect if this is Atlas MongoDB (contains mongodb.net) or local MongoDB
is_atlas = 'mongodb.net' in mongo_url or 'mongodb+srv' in mongo_url

# Configure MongoDB client with proper timeouts and SSL settings
client_options = {
    'serverSelectionTimeoutMS': 30000,  # 30 seconds for server selection
    'connectTimeoutMS': 30000,           # 30 seconds for initial connection
    'socketTimeoutMS': 30000,            # 30 seconds for socket operations
    'maxPoolSize': 50,                   # Connection pool size
    'minPoolSize': 10,                   # Minimum connections to maintain
    'retryWrites': True,                 # Retry write operations
    'retryReads': True,                  # Retry read operations
}

# Add TLS settings only for Atlas MongoDB
if is_atlas:
    client_options.update({
        'tls': True,                         # Enable TLS/SSL for Atlas
        'tlsAllowInvalidCertificates': False # Validate SSL certificates
    })

client = AsyncIOMotorClient(mongo_url, **client_options)
db = client[os.environ['DB_NAME']]

async def next_sequence(name: str) -> int:
    """Atomically increment a named counter — prevents duplicate invoice/job numbers.
    If counter doesn't exist yet, seeds it from actual collection count so numbers
    stay sensible (e.g. INV-000448 after 447 existing sales).
    """
    # Collection name map: counter name -> db collection
    collection_map = {
        "sales": db.sales,
        "services": db.services,
        "registrations": db.registrations,
        "spare_part_bills": db.spare_part_bills,
        "service_bills": db.service_bills,
    }

    # Check if counter already exists
    existing = await db.counters.find_one({"_id": name})
    if not existing:
        # Seed from actual count so we don't start from 0 or a crazy number
        coll = collection_map.get(name)
        seed = await coll.count_documents({}) if coll is not None else 0
        await db.counters.update_one(
            {"_id": name},
            {"$setOnInsert": {"seq": seed}},
            upsert=True
        )

    result = await db.counters.find_one_and_update(
        {"_id": name},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return result["seq"]


# JWT Configuration
SECRET_KEY = os.environ['JWT_SECRET_KEY']
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440  # 24 hours

# Create the main app without a prefix
app = FastAPI(title="Two Wheeler Business Management API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Startup event to verify MongoDB connection
@app.on_event("startup")
async def startup_db_client():
    """Test MongoDB connection and create indexes on startup"""
    try:
        await client.admin.command('ping')
        print("✅ Successfully connected to MongoDB Atlas")
    except Exception as e:
        print(f"❌ Failed to connect to MongoDB: {str(e)}")
        print(f"   Make sure MONGO_URL is correctly configured for Atlas")

    # Create indexes for frequently queried fields
    try:
        await db.vehicles.create_index("chassis_number", sparse=True)
        await db.vehicles.create_index("vehicle_number", sparse=True)
        await db.vehicles.create_index("status")
        await db.customers.create_index("mobile", sparse=True)
        await db.services.create_index("job_card_number", sparse=True)
        await db.services.create_index("status")
        await db.sales.create_index("invoice_number", sparse=True)
        await db.sales.create_index("customer_id")
        await db.spare_parts.create_index("part_number", sparse=True)
        await db.activities.create_index([("created_at", -1)])
        print("✅ MongoDB indexes created")
    except Exception as e:
        print(f"⚠️ Index creation warning: {e}")

    # Fix corrupted sequence counters — reset to actual collection counts
    # This runs every startup but is a no-op if counters are already correct
    try:
        counter_sources = [
            ("sales",            db.sales),
            ("services",         db.services),
            ("registrations",    db.registrations),
            ("spare_part_bills", db.spare_part_bills),
            ("service_bills",    db.service_bills),
        ]
        for counter_name, coll in counter_sources:
            existing = await db.counters.find_one({"_id": counter_name})
            real_count = await coll.count_documents({})
            if existing is None or existing.get("seq", 0) > real_count + 10000:
                # Counter is missing or wildly wrong — reset it to real count
                await db.counters.update_one(
                    {"_id": counter_name},
                    {"$set": {"seq": real_count}},
                    upsert=True
                )
                print(f"✅ Reset {counter_name} counter to {real_count}")
        print("✅ Sequence counters verified")
    except Exception as e:
        print(f"⚠️ Counter reset warning: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    """Close MongoDB connection on shutdown"""
    client.close()
    print("✅ MongoDB connection closed")

# Security
security = HTTPBearer()

# Enums
class UserRole(str, Enum):
    ADMIN = "admin"
    STAFF = "staff"

class VehicleStatus(str, Enum):
    IN_STOCK = "in_stock"
    SOLD = "sold"
    RETURNED = "returned"

class ServiceStatus(str, Enum):
    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"

# Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: EmailStr
    role: UserRole
    full_name: str
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    role: UserRole
    full_name: str

class UserLogin(BaseModel):
    username: str
    password: str

class Customer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: Optional[str] = None
    care_of: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[EmailStr] = None
    address: Optional[str] = None
    vehicle_info: Optional[Dict[str, Any]] = None
    insurance_info: Optional[Dict[str, Any]] = None
    sales_info: Optional[Dict[str, Any]] = None
    customer_type: Optional[str] = "sales"  # "sales", "service", "both"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    name: Optional[str] = None
    care_of: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[EmailStr] = None
    address: Optional[str] = None
    vehicle_info: Optional[Dict[str, Any]] = None
    insurance_info: Optional[Dict[str, Any]] = None
    sales_info: Optional[Dict[str, Any]] = None
    customer_type: Optional[str] = "sales"  # "sales", "service", "both"



class SalesMilestone(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sale_id: str
    customer_name: str
    invoice_number: str
    milestones: Dict[str, Any] = Field(default_factory=dict)
    # milestone keys: customer_docs, invoice_insurance, road_tax, number_plates, plates_attached
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Vehicle(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    brand: Optional[str] = None  # TVS, BAJAJ, HERO, HONDA, TRIUMPH, KTM, SUZUKI, APRILIA, YAMAHA, PIAGGIO, ROYAL ENFIELD
    model: Optional[str] = None
    chassis_number: Optional[str] = None  # Standardized from chassis_no
    engine_number: Optional[str] = None  # Standardized from engine_no
    color: Optional[str] = None
    vehicle_number: Optional[str] = None  # Registration number
    key_number: Optional[str] = None  # Standardized from key_no
    inbound_location: Optional[str] = None
    outbound_location: Optional[str] = None
    status: VehicleStatus = VehicleStatus.IN_STOCK
    page_number: Optional[str] = None
    date_received: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    date_sold: Optional[datetime] = None
    date_returned: Optional[datetime] = None
    customer_id: Optional[str] = None

class VehicleCreate(BaseModel):
    brand: Optional[str] = None
    model: Optional[str] = None
    chassis_number: Optional[str] = None  # Standardized from chassis_no
    engine_number: Optional[str] = None  # Standardized from engine_no
    color: Optional[str] = None
    vehicle_number: Optional[str] = None  # Registration number
    key_number: Optional[str] = None  # Standardized from key_no
    inbound_location: Optional[str] = None
    page_number: Optional[str] = None
    date_received: Optional[datetime] = None

class VehicleUpdate(BaseModel):
    brand: Optional[str] = None
    model: Optional[str] = None
    chassis_number: Optional[str] = None
    engine_number: Optional[str] = None
    color: Optional[str] = None
    vehicle_number: Optional[str] = None
    key_number: Optional[str] = None
    inbound_location: Optional[str] = None
    page_number: Optional[str] = None
    outbound_location: Optional[str] = None
    status: Optional[str] = None
    date_received: Optional[datetime] = None
    date_returned: Optional[str] = None

class Sale(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    customer_id: str
    vehicle_id: Optional[str] = None  # Made optional for imported sales without specific vehicles
    sale_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    amount: float
    payment_method: str
    insurance_details: Optional[Dict[str, Any]] = None
    created_by: str
    source: str = "direct"  # "direct" for manual sales, "import" for imported data
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    # Additional fields for imported sales data (not in create model)
    vehicle_brand: Optional[str] = None
    vehicle_model: Optional[str] = None
    vehicle_color: Optional[str] = None
    vehicle_chassis: Optional[str] = None
    vehicle_engine: Optional[str] = None
    vehicle_registration: Optional[str] = None
    insurance_nominee: Optional[str] = None
    insurance_relation: Optional[str] = None
    insurance_age: Optional[str] = None
    hypothecation: Optional[str] = None
    pending_amount: Optional[float] = None  # None = fully paid, >0 = balance due

class SaleCreate(BaseModel):
    customer_id: str
    vehicle_id: Optional[str] = None  # Made optional for imported sales
    sale_date: Optional[datetime] = None  # Allow updating sale date
    amount: float
    payment_method: str
    insurance_details: Optional[Dict[str, Any]] = None
    source: str = "direct"  # Default to direct sales
    
    # Additional fields for imported sales and editing
    vehicle_brand: Optional[str] = None
    vehicle_model: Optional[str] = None
    vehicle_color: Optional[str] = None
    vehicle_chassis: Optional[str] = None
    vehicle_engine: Optional[str] = None
    vehicle_registration: Optional[str] = None
    insurance_nominee: Optional[str] = None
    insurance_relation: Optional[str] = None
    insurance_age: Optional[str] = None
    hypothecation: Optional[str] = None
    pending_amount: Optional[float] = None  # Balance due after partial payment

class Service(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    job_card_number: str
    customer_id: str
    vehicle_id: Optional[str] = None
    vehicle_number: str  # Registration number (standardized)
    vehicle_brand: Optional[str] = None  # For imported services without vehicle_id
    vehicle_model: Optional[str] = None  # For imported services without vehicle_id
    vehicle_year: Optional[str] = None  # For imported services without vehicle_id
    service_type: str
    description: str
    status: ServiceStatus = ServiceStatus.PENDING
    service_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completion_date: Optional[datetime] = None
    amount: float
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_registration: bool = False  # Flag to identify registration vs job card
    service_number: Optional[str] = None  # User-defined service number
    kms_driven: Optional[int] = None  # Kilometers driven at time of service

# Customer + Vehicle Registration Model (One-time registration)
class Registration(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    registration_number: str  # REG-000001
    customer_id: str
    customer_name: str
    customer_mobile: str
    customer_address: Optional[str] = None
    vehicle_number: str
    vehicle_brand: Optional[str] = None
    vehicle_model: Optional[str] = None
    vehicle_year: Optional[str] = None
    chassis_number: Optional[str] = None
    engine_number: Optional[str] = None
    registration_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RegistrationCreate(BaseModel):
    customer_name: str
    customer_mobile: str
    customer_address: Optional[str] = None
    vehicle_number: str
    vehicle_brand: Optional[str] = None
    vehicle_model: Optional[str] = None
    vehicle_year: Optional[str] = None
    chassis_number: Optional[str] = None
    engine_number: Optional[str] = None

class ServiceCreate(BaseModel):
    customer_id: str
    vehicle_id: Optional[str] = None
    vehicle_number: str  # Registration number (standardized)
    vehicle_brand: Optional[str] = None
    vehicle_model: Optional[str] = None
    vehicle_year: Optional[str] = None
    service_type: str
    description: str
    amount: float
    service_number: Optional[str] = None
    kms_driven: Optional[int] = None
    service_date: Optional[datetime] = None  # Service date (defaults to now if not provided)

class ServiceUpdate(BaseModel):
    customer_id: str
    vehicle_id: Optional[str] = None
    vehicle_number: str  # Registration number (standardized)
    vehicle_brand: Optional[str] = None
    vehicle_model: Optional[str] = None
    vehicle_year: Optional[str] = None
    service_type: str
    description: str
    amount: float
    service_date: Optional[datetime] = None  # Registration date
    service_number: Optional[str] = None
    kms_driven: Optional[int] = None

class SparePart(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    part_number: str
    brand: str
    quantity: int
    unit: str = "Nos"  # Unit of measurement (Nos, Kg, Ltr, etc.)
    unit_price: float
    hsn_sac: Optional[str] = None  # HSN/SAC code
    gst_percentage: float = 18.0  # GST percentage
    compatible_models: Optional[str] = None  # Compatible vehicle models
    low_stock_threshold: int = 5
    supplier: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SparePartCreate(BaseModel):
    name: str
    part_number: str
    brand: str
    quantity: int
    unit: str = "Nos"
    unit_price: float
    hsn_sac: Optional[str] = None
    gst_percentage: float = 18.0
    compatible_models: Optional[str] = None
    low_stock_threshold: int = 5
    supplier: Optional[str] = None

class SparePartBill(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bill_number: str
    customer_id: Optional[str] = None  # For backwards compatibility
    customer_data: Optional[Dict[str, str]] = None  # {name, mobile, vehicle_name, vehicle_number}
    items: List[Dict[str, Any]]  # Detailed GST items with all calculations
    subtotal: float
    total_discount: float
    total_cgst: float
    total_sgst: float
    total_tax: float
    total_amount: float
    bill_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SparePartBillCreate(BaseModel):
    customer_data: Optional[Dict[str, str]] = None  # {name, mobile, vehicle_name, vehicle_number}
    customer_id: Optional[str] = None  # For backwards compatibility
    items: List[Dict[str, Any]]
    subtotal: Optional[float] = 0
    total_discount: Optional[float] = 0
    total_cgst: Optional[float] = 0
    total_sgst: Optional[float] = 0
    total_tax: Optional[float] = 0
    total_amount: Optional[float] = 0

# Service Bill Models
class ServiceBill(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bill_number: str
    job_card_number: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_mobile: Optional[str] = None
    vehicle_number: Optional[str] = None
    vehicle_brand: Optional[str] = None
    vehicle_model: Optional[str] = None
    items: List[Dict[str, Any]]  # Itemized bill items with GST calculations
    subtotal: float = 0
    total_discount: float = 0
    total_cgst: float = 0
    total_sgst: float = 0
    total_tax: float = 0
    total_amount: float = 0
    bill_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "pending"
    created_by: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ServiceBillCreate(BaseModel):
    bill_number: str
    job_card_number: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_mobile: Optional[str] = None
    vehicle_number: Optional[str] = None
    vehicle_brand: Optional[str] = None
    vehicle_model: Optional[str] = None
    items: List[Dict[str, Any]]
    subtotal: float = 0
    total_discount: float = 0
    total_cgst: float = 0
    total_sgst: float = 0
    total_tax: float = 0
    total_amount: float = 0
    bill_date: Optional[str] = None
    status: str = "pending"

# Backup Models
class BackupConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    backup_enabled: bool = True
    backup_time: str = "02:00"  # 24-hour format
    retention_days: int = 30
    compress_backups: bool = True
    email_notifications: bool = False
    email_recipients: List[str] = []
    backup_location: str = "./backups"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BackupJob(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status: str  # running, completed, failed
    start_time: datetime
    end_time: Optional[datetime] = None
    total_records: int = 0
    backup_size_mb: float = 0
    backup_file_path: str = ""
    error_message: Optional[str] = None
    records_backed_up: Dict[str, int] = {}
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BackupJobCreate(BaseModel):
    backup_type: str = "manual"  # manual or scheduled
    export_format: str = "json"  # json, excel

class BackupStats(BaseModel):
    total_backups: int
    successful_backups: int
    failed_backups: int
    last_backup_date: Optional[datetime]
    total_storage_used_mb: float
    oldest_backup_date: Optional[datetime]

# Import Models
class ImportJob(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    file_name: str
    data_type: str  # customers, vehicles, spare_parts, services
    status: str  # processing, completed, failed
    total_records: int = 0
    processed_records: int = 0
    successful_records: int = 0
    failed_records: int = 0
    skipped_records: int = 0  # Records skipped due to duplicates
    errors: List[Dict[str, Any]] = []
    cross_reference_stats: Optional[Dict[str, int]] = {}  # Track linking statistics
    incomplete_records: List[Dict[str, Any]] = []  # Records with missing data
    start_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    end_time: Optional[datetime] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ImportResult(BaseModel):
    job_id: str
    status: str
    message: str
    total_records: int = 0
    successful_records: int = 0
    failed_records: int = 0
    skipped_records: int = 0  # Records skipped due to duplicates
    errors: List[Dict[str, Any]] = []
    cross_reference_stats: Optional[Dict[str, int]] = {}  # Linking statistics
    incomplete_records: List[Dict[str, Any]] = []  # Records needing completion

# Dismissed Service Due Model - track service due records that have been dismissed/deleted
class DismissedServiceDue(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    service_due_key: str  # Unique key: customer_id-vehicle_number
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    vehicle_reg_no: Optional[str] = None
    dismissed_by: str
    dismissed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    reason: Optional[str] = None

# Service Due Base Date Override Model - store custom base dates
class ServiceDueBaseDateOverride(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    service_due_key: str  # Unique key: customer_id-vehicle_number
    custom_base_date: datetime
    updated_by: str
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    notes: Optional[str] = None

class BulkDeleteRequest(BaseModel):
    ids: List[str]
    force_delete: bool = False

# Utility functions
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed_password.encode('utf-8'))

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        
        user = await db.users.find_one({"username": username})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        return User(**user)
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

async def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Verify JWT token and return user data"""
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        
        user = await db.users.find_one({"username": username})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        # Handle both 'id' and '_id' fields for user identification
        user_id = user.get("id") or str(user.get("_id", ""))
        return {"user_id": user_id, "username": user["username"], "role": user.get("role", "user")}
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

def parse_date_flexible(date_str: str) -> datetime:
    """Parse date from various formats"""
    date_str = date_str.strip()
    
    # Try to parse as Excel serial date number
    try:
        excel_date = float(date_str)
        if excel_date > 59:
            excel_date -= 1
        return datetime(1900, 1, 1, tzinfo=timezone.utc) + timedelta(days=excel_date - 2)
    except (ValueError, TypeError):
        pass
    
    # Try various string date formats
    # Format: DD-MMM (03-Mar) - add current year
    if re.match(r'\d{1,2}-[A-Za-z]{3}', date_str):
        date_str = f"{date_str}-{datetime.now().year}"
        return datetime.strptime(date_str, "%d-%b-%Y").replace(tzinfo=timezone.utc)
    # Format: DD/MM/YYYY (15/01/2024)
    elif re.match(r'\d{1,2}/\d{1,2}/\d{4}', date_str):
        return datetime.strptime(date_str, "%d/%m/%Y").replace(tzinfo=timezone.utc)
    # Format: DD-MM-YYYY (15-01-2024)
    elif re.match(r'\d{1,2}-\d{1,2}-\d{4}', date_str):
        return datetime.strptime(date_str, "%d-%m-%Y").replace(tzinfo=timezone.utc)
    # Format: YYYY-MM-DD (2024-01-15) - ISO format
    elif re.match(r'\d{4}-\d{1,2}-\d{1,2}', date_str):
        return datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    # Format: YYYY/MM/DD (2024/01/15)
    elif re.match(r'\d{4}/\d{1,2}/\d{1,2}', date_str):
        return datetime.strptime(date_str, "%Y/%m/%d").replace(tzinfo=timezone.utc)
    # Format: DD MMM YYYY (15 Jan 2024)
    elif re.match(r'\d{1,2}\s+[A-Za-z]{3}\s+\d{4}', date_str):
        return datetime.strptime(date_str, "%d %b %Y").replace(tzinfo=timezone.utc)
    # Format: MMM DD, YYYY (Jan 15, 2024)
    elif re.match(r'[A-Za-z]{3}\s+\d{1,2},?\s+\d{4}', date_str):
        return datetime.strptime(date_str.replace(',', ''), "%b %d %Y").replace(tzinfo=timezone.utc)
    else:
        # Try generic parser as last resort
        try:
            from dateutil import parser
            return parser.parse(date_str).replace(tzinfo=timezone.utc)
        except (ValueError, TypeError, OverflowError):
            return datetime.now(timezone.utc)

def safe_str(value) -> str:
    """Safely convert a value to string, handling NaN, None, and floats from Excel/pandas"""
    import math
    if value is None:
        return ''
    if isinstance(value, float):
        # Check for NaN
        if math.isnan(value):
            return ''
        # If it's a whole number, convert without decimal
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, (int, bool)):
        return str(value)
    if isinstance(value, str):
        return value.strip()
    return str(value).strip() if value else ''

# Health check endpoints for Kubernetes (both root and /api paths for compatibility)
@app.get("/health")
async def health_check():
    """Health check endpoint for Kubernetes liveness probe"""
    return {"status": "healthy"}

@app.get("/ready")
async def readiness_check():
    """Readiness check endpoint for Kubernetes readiness probe"""
    try:
        # Test database connection
        await db.command("ping")
        return {"status": "ready", "database": "connected"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database not ready: {str(e)}")

# Also add health checks under /api prefix for compatibility
@api_router.get("/health")
async def api_health_check():
    """Health check endpoint at /api/health"""
    return {"status": "healthy"}

@api_router.get("/ready")
async def api_readiness_check():
    """Readiness check endpoint at /api/ready"""
    try:
        # Test database connection
        await db.command("ping")
        return {"status": "ready", "database": "connected"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database not ready: {str(e)}")

# Test endpoint
@api_router.get("/")
async def root():
    return {"message": "Two Wheeler Business Management API is running"}

# Dismissed Service Due endpoints
@api_router.get("/dismissed-service-due")
async def get_dismissed_service_due(current_user: User = Depends(get_current_user)):
    """Get all dismissed service due records"""
    dismissed = await db.dismissed_service_due.find({}, {"_id": 0}).to_list(10000)
    return dismissed

@api_router.post("/dismissed-service-due")
async def dismiss_service_due(data: dict, current_user: User = Depends(get_current_user)):
    """Dismiss a single service due record"""
    dismissed = DismissedServiceDue(
        service_due_key=data.get("service_due_key"),
        customer_id=data.get("customer_id"),
        customer_name=data.get("customer_name"),
        vehicle_reg_no=data.get("vehicle_reg_no"),
        dismissed_by=current_user.id,
        reason=data.get("reason", "Manually dismissed")
    )
    await db.dismissed_service_due.insert_one(dismissed.dict())
    return {"message": "Service due record dismissed successfully", "id": dismissed.id}

@api_router.post("/dismissed-service-due/bulk")
async def bulk_dismiss_service_due(data: dict, current_user: User = Depends(get_current_user)):
    """Bulk dismiss service due records"""
    items = data.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="No items provided for bulk dismiss")
    
    dismissed_count = 0
    for item in items:
        dismissed = DismissedServiceDue(
            service_due_key=item.get("service_due_key"),
            customer_id=item.get("customer_id"),
            customer_name=item.get("customer_name"),
            vehicle_reg_no=item.get("vehicle_reg_no"),
            dismissed_by=current_user.id,
            reason=item.get("reason", "Bulk dismissed")
        )
        await db.dismissed_service_due.insert_one(dismissed.dict())
        dismissed_count += 1
    
    return {"message": f"Successfully dismissed {dismissed_count} service due records", "count": dismissed_count}

@api_router.delete("/dismissed-service-due/{key}")
async def restore_service_due(key: str, current_user: User = Depends(get_current_user)):
    """Restore a dismissed service due record (remove from dismissed list)"""
    result = await db.dismissed_service_due.delete_one({"service_due_key": key})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dismissed record not found")
    return {"message": "Service due record restored successfully"}

# Service Due Base Date Override endpoints
@api_router.get("/service-due-base-date")
async def get_base_date_overrides(current_user: User = Depends(get_current_user)):
    """Get all base date overrides"""
    overrides = await db.service_due_base_date_overrides.find({}, {"_id": 0}).to_list(10000)
    return overrides

@api_router.post("/service-due-base-date")
async def set_base_date_override(data: dict, current_user: User = Depends(get_current_user)):
    """Set or update a base date override for a service due record"""
    service_due_key = data.get("service_due_key")
    custom_base_date = data.get("custom_base_date")
    
    if not service_due_key or not custom_base_date:
        raise HTTPException(status_code=400, detail="service_due_key and custom_base_date are required")
    
    # Parse the date
    try:
        if isinstance(custom_base_date, str):
            parsed_date = datetime.fromisoformat(custom_base_date.replace('Z', '+00:00'))
        else:
            parsed_date = custom_base_date
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    
    # Check if override already exists
    existing = await db.service_due_base_date_overrides.find_one({"service_due_key": service_due_key})
    
    if existing:
        # Update existing override
        await db.service_due_base_date_overrides.update_one(
            {"service_due_key": service_due_key},
            {"$set": {
                "custom_base_date": parsed_date,
                "updated_by": current_user.id,
                "updated_at": datetime.now(timezone.utc),
                "notes": data.get("notes")
            }}
        )
        return {"message": "Base date override updated successfully", "service_due_key": service_due_key}
    else:
        # Create new override
        override = ServiceDueBaseDateOverride(
            service_due_key=service_due_key,
            custom_base_date=parsed_date,
            updated_by=current_user.id,
            notes=data.get("notes")
        )
        await db.service_due_base_date_overrides.insert_one(override.dict())
        return {"message": "Base date override created successfully", "id": override.id}

@api_router.delete("/service-due-base-date/{key}")
async def delete_base_date_override(key: str, current_user: User = Depends(get_current_user)):
    """Delete a base date override (revert to calculated date)"""
    result = await db.service_due_base_date_overrides.delete_one({"service_due_key": key})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Base date override not found")
    return {"message": "Base date override removed successfully"}

# Authentication endpoints
@api_router.post("/auth/register")
async def register_user(user_data: UserCreate):
    # Check if user exists
    existing_user = await db.users.find_one({"$or": [{"username": user_data.username}, {"email": user_data.email}]})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username or email already registered")
    
    # Hash password
    hashed_password = hash_password(user_data.password)
    
    # Create user document for database
    user_dict = user_data.dict()
    user_dict.pop('password')
    user_dict['hashed_password'] = hashed_password
    user_dict['id'] = str(uuid.uuid4())
    user_dict['is_active'] = True
    user_dict['created_at'] = datetime.now(timezone.utc)
    
    await db.users.insert_one(user_dict)
    return {"message": "User registered successfully", "user_id": user_dict['id']}

@api_router.post("/auth/login")
async def login_user(user_credentials: UserLogin):
    user = await db.users.find_one({"username": user_credentials.username})
    if not user or not verify_password(user_credentials.password, user['hashed_password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user['is_active']:
        raise HTTPException(status_code=401, detail="User account is inactive")
    
    access_token = create_access_token(data={"sub": user['username'], "role": user['role']})
    return {"access_token": access_token, "token_type": "bearer", "user": User(**user)}

@api_router.get("/auth/me")
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    return current_user

# Customer endpoints
@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, current_user: User = Depends(get_current_user)):
    # Check for duplicate mobile number
    if customer_data.mobile and await check_customer_duplicate(customer_data.mobile):
        raise HTTPException(status_code=400, detail=f"Customer with mobile number '{customer_data.mobile}' already exists")
    
    customer = Customer(**customer_data.dict())
    await db.customers.insert_one(customer.dict())
    return customer


@api_router.get("/customers/by-mobile/{mobile}")
async def get_customer_by_mobile(mobile: str, current_user: User = Depends(get_current_user)):
    """Look up a customer by exact mobile number — used by invoice creation to avoid duplicates"""
    customer = await db.customers.find_one({"mobile": mobile})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    customer.pop("_id", None)
    return customer

@api_router.get("/customers")
async def get_customers(
    page: int = 1,
    limit: int = 100,
    sort: str = "created_at",
    order: str = "desc",
    customer_type: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Validate sort field
    valid_sort_fields = ["name", "mobile", "created_at", "total_purchases"]
    if sort not in valid_sort_fields:
        raise HTTPException(status_code=400, detail=f"Invalid sort field. Must be one of: {', '.join(valid_sort_fields)}")
    
    # Validate order
    if order not in ["asc", "desc"]:
        raise HTTPException(status_code=400, detail="Invalid order. Must be 'asc' or 'desc'")
    
    # Calculate skip
    skip = (page - 1) * limit
    
    # Build sort criteria
    sort_direction = 1 if order == "asc" else -1

    # Build base query filter
    base_filter = {}
    if customer_type:
        if customer_type == "sales":
            base_filter["$or"] = [{"customer_type": "sales"}, {"customer_type": {"$exists": False}}, {"customer_type": None}]
        elif customer_type == "service":
            base_filter["customer_type"] = "service"
        elif customer_type == "both":
            base_filter["customer_type"] = "both"
    
    # For total_purchases, we need to aggregate
    if sort == "total_purchases":
        # Aggregate to calculate total purchases
        pipeline = [
            {"$match": base_filter},
            {
                "$lookup": {
                    "from": "sales",
                    "localField": "id",
                    "foreignField": "customer_id",
                    "as": "sales"
                }
            },
            {
                "$addFields": {
                    "total_purchases": {"$size": "$sales"}
                }
            },
            {"$sort": {"total_purchases": sort_direction}},
            {"$skip": skip},
            {"$limit": limit}
        ]
        customers = await db.customers.aggregate(pipeline).to_list(None)
        
        # Get total count
        total = await db.customers.count_documents(base_filter)
    else:
        # Regular sort
        customers = await db.customers.find(base_filter).sort(sort, sort_direction).skip(skip).limit(limit).to_list(None)
        total = await db.customers.count_documents(base_filter)
    
    # Convert ObjectId to string for JSON serialization
    for customer in customers:
        if '_id' in customer:
            customer['_id'] = str(customer['_id'])
    
    # Calculate total pages
    total_pages = (total + limit - 1) // limit
    
    return {
        "data": customers,
        "meta": {
            "total": total,
            "page": page,
            "limit": limit,
            "totalPages": total_pages
        }
    }

@api_router.get("/customers/{customer_id}")
async def get_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    # Convert ObjectId to string for JSON serialization
    if '_id' in customer:
        customer['_id'] = str(customer['_id'])
    return customer

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer_data: CustomerCreate, current_user: User = Depends(get_current_user)):
    # Check if customer exists
    existing_customer = await db.customers.find_one({"id": customer_id})
    if not existing_customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Prepare update data by merging with existing data to preserve nested fields
    update_data = customer_data.dict(exclude_unset=True)  # Only include fields that were explicitly set
    
    # Always preserve these fields
    update_data["id"] = customer_id  # Keep the original ID
    update_data["created_at"] = existing_customer["created_at"]  # Keep original creation date
    
    # Merge nested fields to preserve existing data
    # Handle vehicle_info preservation
    if "vehicle_info" in update_data:
        existing_vehicle_info = existing_customer.get("vehicle_info", {})
        if existing_vehicle_info and update_data["vehicle_info"]:
            # Merge existing vehicle_info with new vehicle_info
            merged_vehicle_info = {**existing_vehicle_info, **update_data["vehicle_info"]}
            update_data["vehicle_info"] = merged_vehicle_info
    else:
        # Preserve existing vehicle_info if not included in update
        if "vehicle_info" in existing_customer:
            update_data["vehicle_info"] = existing_customer["vehicle_info"]
    
    # Handle insurance_info preservation
    if "insurance_info" in update_data:
        existing_insurance_info = existing_customer.get("insurance_info", {})
        if existing_insurance_info and update_data["insurance_info"]:
            # Merge existing insurance_info with new insurance_info
            merged_insurance_info = {**existing_insurance_info, **update_data["insurance_info"]}
            update_data["insurance_info"] = merged_insurance_info
    else:
        # Preserve existing insurance_info if not included in update
        if "insurance_info" in existing_customer:
            update_data["insurance_info"] = existing_customer["insurance_info"]
    
    # Handle sales_info preservation
    if "sales_info" in update_data:
        existing_sales_info = existing_customer.get("sales_info", {})
        if existing_sales_info and update_data["sales_info"]:
            # Merge existing sales_info with new sales_info
            merged_sales_info = {**existing_sales_info, **update_data["sales_info"]}
            update_data["sales_info"] = merged_sales_info
    else:
        # Preserve existing sales_info if not included in update
        if "sales_info" in existing_customer:
            update_data["sales_info"] = existing_customer["sales_info"]
    
    # Create the complete updated customer data by merging with existing
    complete_update_data = {**existing_customer, **update_data}
    
    updated_customer = Customer(**complete_update_data)
    await db.customers.replace_one({"id": customer_id}, updated_customer.dict())
    return updated_customer

# Vehicle endpoints
@api_router.post("/vehicles", response_model=Vehicle)
async def create_vehicle(vehicle_data: VehicleCreate, current_user: User = Depends(get_current_user)):
    # Check for duplicate chassis number
    if vehicle_data.chassis_number and await check_vehicle_duplicate(vehicle_data.chassis_number):
        raise HTTPException(status_code=400, detail=f"Vehicle with chassis number '{vehicle_data.chassis_number}' already exists")
    
    # Create vehicle dict and handle date_received
    vehicle_dict = vehicle_data.dict()
    if vehicle_dict.get('date_received') is None:
        vehicle_dict['date_received'] = datetime.now(timezone.utc)
    
    vehicle = Vehicle(**vehicle_dict)
    await db.vehicles.insert_one(vehicle.dict())
    
    # Create activity notification
    try:
        await create_activity(ActivityCreate(
            type=ActivityType.VEHICLE_ADDED,
            title="New vehicle added to stock",
            description=f"{vehicle_data.brand} {vehicle_data.model} - {vehicle_data.chassis_number}",
            icon="info",
            metadata={"vehicle_id": vehicle.id}
        ))
    except Exception as e:
        logger.warning(f"Failed to create activity for vehicle addition: {e}")
    
    return vehicle

@api_router.get("/vehicles", response_model=List[Vehicle])
async def get_vehicles(brand: Optional[str] = None, status: Optional[VehicleStatus] = None, current_user: User = Depends(get_current_user)):
    filter_dict = {}
    if brand:
        filter_dict["brand"] = brand
    if status:
        filter_dict["status"] = status
    
    vehicles = await db.vehicles.find(filter_dict).to_list(1000)
    return [Vehicle(**vehicle) for vehicle in vehicles]

@api_router.get("/vehicles/brands")
async def get_vehicle_brands(current_user: User = Depends(get_current_user)):
    brands = ["TVS", "BAJAJ", "HERO", "HONDA", "TRIUMPH", "KTM", "SUZUKI", "APRILIA", "YAMAHA", "PIAGGIO", "ROYAL ENFIELD"]
    return brands

@api_router.put("/vehicles/{vehicle_id}", response_model=Vehicle)
async def update_vehicle(vehicle_id: str, vehicle_data: VehicleUpdate, current_user: User = Depends(get_current_user)):
    # Check if vehicle exists
    existing_vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not existing_vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Update vehicle data - only update fields that are provided
    update_data = {k: v for k, v in vehicle_data.dict().items() if v is not None}
    update_data["id"] = vehicle_id  # Keep the original ID
    
    # Handle date_returned if provided
    if "date_returned" in update_data and update_data["date_returned"]:
        try:
            from dateutil import parser as date_parser
            update_data["date_returned"] = date_parser.parse(update_data["date_returned"])
        except (ValueError, TypeError):
            pass
    
    # Merge existing data with updates
    merged_data = {**existing_vehicle, **update_data}
    
    updated_vehicle = Vehicle(**merged_data)
    await db.vehicles.replace_one({"id": vehicle_id}, updated_vehicle.dict())
    return updated_vehicle

@api_router.get("/vehicles/{vehicle_id}", response_model=Vehicle)
async def get_vehicle(vehicle_id: str, current_user: User = Depends(get_current_user)):
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    return Vehicle(**vehicle)

class VehicleStatusUpdate(BaseModel):
    status: str
    return_date: Optional[str] = None
    outbound_location: Optional[str] = None

@api_router.put("/vehicles/{vehicle_id}/status")
async def update_vehicle_status(vehicle_id: str, status_data: VehicleStatusUpdate, current_user: User = Depends(get_current_user)):
    """Update vehicle status with optional return date"""
    # Check if vehicle exists
    existing_vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not existing_vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Validate status
    new_status = status_data.status
    if not new_status:
        raise HTTPException(status_code=400, detail="Status is required")
    
    # Validate status value
    valid_statuses = ["in_stock", "sold", "returned"]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    # Prepare update data
    update_data = {"status": new_status}
    
    # Handle status-specific updates
    if new_status == "sold":
        update_data["date_sold"] = datetime.now(timezone.utc)
        # Clear return date if previously returned
        update_data["date_returned"] = None
    elif new_status == "returned":
        # Set return date
        return_date = status_data.return_date
        if return_date:
            try:
                # Parse the return date if provided
                update_data["date_returned"] = datetime.fromisoformat(return_date.replace('Z', '+00:00'))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid return_date format. Use ISO format (YYYY-MM-DDTHH:MM:SSZ)")
        else:
            update_data["date_returned"] = datetime.now(timezone.utc)
        
        # Set outbound location if provided
        outbound_location = status_data.outbound_location
        if outbound_location:
            update_data["outbound_location"] = outbound_location
    elif new_status == "in_stock":
        # Clear sold/returned dates when back in stock
        update_data["date_sold"] = None
        update_data["date_returned"] = None
        update_data["customer_id"] = None
        update_data["outbound_location"] = None
    
    # Update the vehicle
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": update_data})
    
    # Return updated vehicle
    updated_vehicle = await db.vehicles.find_one({"id": vehicle_id})
    return Vehicle(**updated_vehicle)

@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, current_user: User = Depends(get_current_user)):
    # Check if vehicle exists
    existing_vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not existing_vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Check if vehicle has associated sales records
    sales_count = await db.sales.count_documents({"vehicle_id": vehicle_id})
    if sales_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete vehicle. Vehicle has {sales_count} associated sales record(s). Please delete sales records first.")
    
    # Check if vehicle has associated service records
    services_count = await db.services.count_documents({"vehicle_id": vehicle_id})
    if services_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete vehicle. Vehicle has {services_count} associated service record(s). Please delete service records first.")
    
    # Delete the vehicle
    result = await db.vehicles.delete_one({"id": vehicle_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    return {"message": "Vehicle deleted successfully", "deleted_vehicle_id": vehicle_id}

@api_router.delete("/vehicles")
async def bulk_delete_vehicles(request: BulkDeleteRequest, current_user: User = Depends(get_current_user)):
    """Bulk delete vehicles with optional force delete (cascade)"""
    if not request.ids:
        raise HTTPException(status_code=400, detail="No vehicle IDs provided")
    
    deleted = []
    failed = []
    cascade_stats = {"sales": 0, "services": 0}
    
    for vehicle_id in request.ids:
        try:
            # Check if vehicle exists
            existing_vehicle = await db.vehicles.find_one({"id": vehicle_id})
            if not existing_vehicle:
                failed.append({"id": vehicle_id, "error": "Vehicle not found"})
                continue
            
            # Check for associated records
            sales_count = await db.sales.count_documents({"vehicle_id": vehicle_id})
            services_count = await db.services.count_documents({"vehicle_id": vehicle_id})
            
            # If force delete is enabled, delete associated records first
            if request.force_delete:
                # Delete associated sales
                if sales_count > 0:
                    sales_result = await db.sales.delete_many({"vehicle_id": vehicle_id})
                    cascade_stats["sales"] += sales_result.deleted_count
                
                # Delete associated services
                if services_count > 0:
                    services_result = await db.services.delete_many({"vehicle_id": vehicle_id})
                    cascade_stats["services"] += services_result.deleted_count
            else:
                # Normal delete - check for restrictions
                if sales_count > 0:
                    failed.append({"id": vehicle_id, "error": f"Vehicle has {sales_count} associated sales record(s)"})
                    continue
                
                if services_count > 0:
                    failed.append({"id": vehicle_id, "error": f"Vehicle has {services_count} associated service record(s)"})
                    continue
            
            # Delete the vehicle
            result = await db.vehicles.delete_one({"id": vehicle_id})
            if result.deleted_count > 0:
                deleted.append(vehicle_id)
            else:
                failed.append({"id": vehicle_id, "error": "Failed to delete"})
        except Exception as e:
            failed.append({"id": vehicle_id, "error": str(e)})
    
    response = {
        "deleted": len(deleted),
        "deleted_ids": deleted,
        "failed": failed
    }
    
    # Include cascade statistics if force delete was used
    if request.force_delete:
        response["cascade_deleted"] = cascade_stats
    
    return response

# Sales endpoints
@api_router.post("/sales", response_model=Sale)
async def create_sale(sale_data: SaleCreate, current_user: User = Depends(get_current_user)):
    # Validate customer exists
    customer = await db.customers.find_one({"id": sale_data.customer_id})
    if not customer:
        raise HTTPException(status_code=400, detail="Customer not found")
    
    # Validate vehicle if provided
    if sale_data.vehicle_id:
        vehicle = await db.vehicles.find_one({"id": sale_data.vehicle_id})
        if not vehicle:
            raise HTTPException(status_code=400, detail="Vehicle not found")
        if vehicle['status'] != VehicleStatus.IN_STOCK:
            raise HTTPException(status_code=400, detail="Vehicle not available for sale")
    
    # Validate amount is positive
    if sale_data.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    
    # Validate payment method
    valid_payment_methods = ["Cash", "Card", "UPI", "Bank Transfer", "Cheque", "Finance"]
    if sale_data.payment_method not in valid_payment_methods:
        raise HTTPException(status_code=400, detail=f"Invalid payment method. Must be one of: {', '.join(valid_payment_methods)}")
    
    # Generate invoice number (atomic — prevents duplicates under concurrent load)
    seq = await next_sequence("sales")
    invoice_number = f"INV-{seq:06d}"
    
    sale_dict = sale_data.dict()
    sale_dict['invoice_number'] = invoice_number
    sale_dict['created_by'] = current_user.id
    
    # Remove None sale_date to allow default factory to work
    if sale_dict.get('sale_date') is None:
        sale_dict.pop('sale_date', None)
    
    sale = Sale(**sale_dict)
    
    # Update vehicle status if vehicle_id provided
    if sale_data.vehicle_id:
        await db.vehicles.update_one(
            {"id": sale_data.vehicle_id},
            {"$set": {"status": VehicleStatus.SOLD, "customer_id": sale_data.customer_id, "date_sold": datetime.now(timezone.utc)}}
        )
    
    await db.sales.insert_one(sale.dict())
    
    # Create activity notification
    try:
        vehicle_info = ""
        if sale_data.vehicle_id:
            vehicle = await db.vehicles.find_one({"id": sale_data.vehicle_id}, {"_id": 0})
            if vehicle:
                vehicle_info = f" - {vehicle.get('brand', '')} {vehicle.get('model', '')}"
        
        await create_activity(ActivityCreate(
            type=ActivityType.SALE_CREATED,
            title="New sale recorded",
            description=f"{customer.get('name', 'Unknown')} - Invoice {invoice_number}{vehicle_info}",
            icon="success",
            metadata={"sale_id": sale.id, "customer_id": sale_data.customer_id, "invoice_number": invoice_number}
        ))
    except Exception as e:
        logger.warning(f"Failed to create activity for sale: {e}")
    
    return sale

@api_router.get("/sales", response_model=List[Sale])
async def get_sales(current_user: User = Depends(get_current_user)):
    sales = await db.sales.find().to_list(1000)
    return [Sale(**sale) for sale in sales]

@api_router.get("/sales/summary/chart")
async def get_sales_summary(
    granularity: str = "monthly",
    years: int = 5,
    current_user: User = Depends(get_current_user)
):
    """Get sales summary for chart - monthly or yearly"""
    if granularity not in ["monthly", "yearly"]:
        raise HTTPException(status_code=400, detail="Granularity must be 'monthly' or 'yearly'")
    
    from datetime import datetime, timezone
    
    if granularity == "monthly":
        # Get last 12 months of data
        pipeline = [
            {
                "$addFields": {
                    "month": {"$month": "$sale_date"},
                    "year": {"$year": "$sale_date"}
                }
            },
            {
                "$group": {
                    "_id": {
                        "year": "$year",
                        "month": "$month"
                    },
                    "total_amount": {"$sum": "$amount"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"_id.year": 1, "_id.month": 1}},
            {"$limit": 12}
        ]
        
        results = await db.sales.aggregate(pipeline).to_list(None)
        
        labels = []
        values = []
        for result in results:
            month_name = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"][result["_id"]["month"] - 1]
            labels.append(f"{month_name} {result['_id']['year']}")
            values.append(result["total_amount"])
        
        return {"labels": labels, "values": values, "granularity": "monthly"}
    
    else:  # yearly
        # Get last N years of data
        pipeline = [
            {
                "$addFields": {
                    "year": {"$year": "$sale_date"}
                }
            },
            {
                "$group": {
                    "_id": "$year",
                    "total_amount": {"$sum": "$amount"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"_id": 1}},
            {"$limit": years}
        ]
        
        results = await db.sales.aggregate(pipeline).to_list(None)
        
        labels = [str(result["_id"]) for result in results]
        values = [result["total_amount"] for result in results]
        
        return {"labels": labels, "values": values, "granularity": "yearly"}

@api_router.get("/sales/{sale_id}", response_model=Sale)
async def get_sale(sale_id: str, current_user: User = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    return Sale(**sale)

@api_router.put("/sales/{sale_id}", response_model=Sale)
async def update_sale(sale_id: str, sale_data: SaleCreate, current_user: User = Depends(get_current_user)):
    # Check if sale exists
    existing_sale = await db.sales.find_one({"id": sale_id})
    if not existing_sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    # Update sale data
    update_data = sale_data.dict(exclude_unset=True)
    update_data["id"] = sale_id  # Keep the original ID
    update_data["invoice_number"] = existing_sale["invoice_number"]  # Keep original invoice number
    update_data["created_by"] = existing_sale["created_by"]  # Keep original creator
    update_data["created_at"] = existing_sale["created_at"]  # Keep original creation date
    
    # If sale_date is provided, use it; otherwise keep existing
    if "sale_date" not in update_data or update_data["sale_date"] is None:
        update_data["sale_date"] = existing_sale["sale_date"]
    
    # Merge with existing data to preserve all fields
    merged_data = {**existing_sale, **update_data}
    
    updated_sale = Sale(**merged_data)
    await db.sales.replace_one({"id": sale_id}, updated_sale.dict())
    return updated_sale

@api_router.patch("/sales/{sale_id}/payment")
async def record_sale_payment(sale_id: str, payment: dict, current_user: User = Depends(get_current_user)):
    """Record a partial or full payment against a sale. Body: { amount_paid: float, note: str? }"""
    existing_sale = await db.sales.find_one({"id": sale_id})
    if not existing_sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    amount_paid = float(payment.get("amount_paid", 0))
    if amount_paid <= 0:
        raise HTTPException(status_code=400, detail="amount_paid must be greater than 0")
    
    current_pending = existing_sale.get("pending_amount")
    total_amount = existing_sale.get("amount", 0)
    
    # If pending_amount was never set, treat full amount as pending
    if current_pending is None:
        current_pending = total_amount
    
    new_pending = max(0, current_pending - amount_paid)
    
    await db.sales.update_one(
        {"id": sale_id},
        {"$set": {"pending_amount": new_pending if new_pending > 0 else 0}}
    )
    
    updated = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    return {"message": "Payment recorded", "pending_amount": new_pending, "sale": updated}

@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str, current_user: User = Depends(get_current_user)):
    # Check if sale exists
    existing_sale = await db.sales.find_one({"id": sale_id})
    if not existing_sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    # If vehicle is associated, reset its status to in_stock
    if existing_sale.get("vehicle_id"):
        await db.vehicles.update_one(
            {"id": existing_sale["vehicle_id"]},
            {"$set": {"status": VehicleStatus.IN_STOCK}, "$unset": {"customer_id": "", "date_sold": ""}}
        )
    
    # Delete the sale
    result = await db.sales.delete_one({"id": sale_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    return {"message": "Sale deleted successfully", "deleted_sale_id": sale_id}

@api_router.delete("/sales")
async def bulk_delete_sales(request: BulkDeleteRequest, current_user: User = Depends(get_current_user)):
    """Bulk delete sales/invoices"""
    if not request.ids:
        raise HTTPException(status_code=400, detail="No sale IDs provided")
    
    deleted = []
    failed = []
    
    for sale_id in request.ids:
        try:
            # Check if sale exists
            existing_sale = await db.sales.find_one({"id": sale_id})
            if not existing_sale:
                failed.append({"id": sale_id, "error": "Sale not found"})
                continue
            
            # If vehicle is associated, reset its status to in_stock
            if existing_sale.get("vehicle_id"):
                await db.vehicles.update_one(
                    {"id": existing_sale["vehicle_id"]},
                    {"$set": {"status": VehicleStatus.IN_STOCK}, "$unset": {"customer_id": "", "date_sold": ""}}
                )
            
            # Delete the sale
            result = await db.sales.delete_one({"id": sale_id})
            if result.deleted_count > 0:
                deleted.append(sale_id)
            else:
                failed.append({"id": sale_id, "error": "Failed to delete"})
        except Exception as e:
            failed.append({"id": sale_id, "error": str(e)})
    
    return {
        "deleted": len(deleted),
        "deleted_ids": deleted,
        "failed": failed
    }

# Registration endpoints (One-time customer + vehicle registration)
@api_router.post("/registrations", response_model=Registration)
async def create_registration(reg_data: RegistrationCreate, current_user: User = Depends(get_current_user)):
    # Check if customer already exists by mobile
    existing_customer = await db.customers.find_one({"mobile": reg_data.customer_mobile})
    
    if existing_customer:
        customer_id = existing_customer["id"]
        # Update customer name if different; also ensure customer_type includes service
        updates = {}
        if existing_customer.get("name") != reg_data.customer_name:
            updates["name"] = reg_data.customer_name
        existing_type = existing_customer.get("customer_type", "sales")
        if existing_type == "sales":
            updates["customer_type"] = "both"
        elif not existing_type:
            updates["customer_type"] = "service"
        if updates:
            await db.customers.update_one({"id": customer_id}, {"$set": updates})
    else:
        # Create new customer
        customer_id = str(uuid.uuid4())
        customer = {
            "id": customer_id,
            "name": reg_data.customer_name,
            "mobile": reg_data.customer_mobile,
            "address": reg_data.customer_address or "",
            "customer_type": "service",
            "created_at": datetime.now(timezone.utc)
        }
        await db.customers.insert_one(customer)
    
    # Generate registration number
    seq = await next_sequence("registrations")
    registration_number = f"REG-{seq:06d}"
    
    registration = Registration(
        registration_number=registration_number,
        customer_id=customer_id,
        customer_name=reg_data.customer_name,
        customer_mobile=reg_data.customer_mobile,
        customer_address=reg_data.customer_address,
        vehicle_number=reg_data.vehicle_number,
        vehicle_brand=reg_data.vehicle_brand,
        vehicle_model=reg_data.vehicle_model,
        vehicle_year=reg_data.vehicle_year,
        chassis_number=reg_data.chassis_number,
        engine_number=reg_data.engine_number,
        created_by=current_user.id
    )
    
    await db.registrations.insert_one(registration.dict())
    return registration

@api_router.get("/registrations")
async def get_registrations(current_user: User = Depends(get_current_user)):
    registrations = await db.registrations.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return registrations

@api_router.get("/registrations/{reg_id}")
async def get_registration(reg_id: str, current_user: User = Depends(get_current_user)):
    registration = await db.registrations.find_one({"id": reg_id}, {"_id": 0})
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    return registration

@api_router.put("/registrations/{reg_id}")
async def update_registration(reg_id: str, reg_data: RegistrationCreate, current_user: User = Depends(get_current_user)):
    existing = await db.registrations.find_one({"id": reg_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Registration not found")
    
    update_data = reg_data.dict()
    update_data["id"] = reg_id
    update_data["registration_number"] = existing["registration_number"]
    update_data["customer_id"] = existing["customer_id"]
    update_data["created_by"] = existing["created_by"]
    update_data["created_at"] = existing["created_at"]
    update_data["registration_date"] = existing.get("registration_date", datetime.now(timezone.utc))
    
    await db.registrations.replace_one({"id": reg_id}, update_data)
    return update_data

@api_router.delete("/registrations/{reg_id}")
async def delete_registration(reg_id: str, current_user: User = Depends(get_current_user)):
    existing = await db.registrations.find_one({"id": reg_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Registration not found")
    
    await db.registrations.delete_one({"id": reg_id})
    return {"message": "Registration deleted successfully"}

# Service endpoints
@api_router.post("/services", response_model=Service)
async def create_service(service_data: ServiceCreate, current_user: User = Depends(get_current_user)):
    # Generate job card number (atomic)
    seq = await next_sequence("services")
    job_card_number = f"JOB-{seq:06d}"
    
    service_dict = service_data.dict()
    service_dict['job_card_number'] = job_card_number
    service_dict['created_by'] = current_user.id
    
    # If service_date is provided, use it; otherwise default will be applied by Service model
    if service_dict.get('service_date') is None:
        service_dict['service_date'] = datetime.now(timezone.utc)
    
    service = Service(**service_dict)
    await db.services.insert_one(service.dict())

    # Tag this customer as a service customer (or "both" if they also have sales)
    if service_dict.get('customer_id'):
        existing = await db.customers.find_one({"id": service_dict['customer_id']})
        if existing:
            current_type = existing.get("customer_type", "sales")
            if current_type == "sales" or not current_type:
                await db.customers.update_one(
                    {"id": service_dict['customer_id']},
                    {"$set": {"customer_type": "both"}}
                )
            elif current_type not in ("service", "both"):
                await db.customers.update_one(
                    {"id": service_dict['customer_id']},
                    {"$set": {"customer_type": "service"}}
                )

    return service

@api_router.get("/services", response_model=List[Service])
async def get_services(status: Optional[ServiceStatus] = None, current_user: User = Depends(get_current_user)):
    filter_dict = {}
    if status:
        filter_dict["status"] = status
    
    services = await db.services.find(filter_dict).to_list(1000)
    return [Service(**service) for service in services]

@api_router.get("/services/{service_id}", response_model=Service)
async def get_service(service_id: str, current_user: User = Depends(get_current_user)):
    service = await db.services.find_one({"id": service_id})
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    return Service(**service)

@api_router.put("/services/{service_id}", response_model=Service)
async def update_service(service_id: str, service_data: ServiceUpdate, current_user: User = Depends(get_current_user)):
    # Check if service exists
    existing_service = await db.services.find_one({"id": service_id})
    if not existing_service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    # Update service data
    update_data = service_data.dict()
    update_data["id"] = service_id  # Keep the original ID
    update_data["job_card_number"] = existing_service["job_card_number"]  # Keep original job card number
    update_data["created_by"] = existing_service["created_by"]  # Keep original creator
    update_data["created_at"] = existing_service["created_at"]  # Keep original creation date
    update_data["status"] = existing_service.get("status", ServiceStatus.PENDING)  # Keep current status
    update_data["completion_date"] = existing_service.get("completion_date")  # Keep completion date if exists
    
    # If service_date not provided, keep the existing one
    if update_data.get("service_date") is None:
        update_data["service_date"] = existing_service.get("service_date")
    
    updated_service = Service(**update_data)
    await db.services.replace_one({"id": service_id}, updated_service.dict())
    return updated_service

@api_router.put("/services/{service_id}/status")
async def update_service_status(service_id: str, status_data: dict, current_user: User = Depends(get_current_user)):
    status = status_data.get("status")
    if not status:
        raise HTTPException(status_code=400, detail="Status is required")
    
    update_data = {"status": status}
    if status == ServiceStatus.COMPLETED:
        update_data["completion_date"] = datetime.now(timezone.utc)
    
    await db.services.update_one({"id": service_id}, {"$set": update_data})
    
    # Create activity notification for completed services
    if status == ServiceStatus.COMPLETED:
        try:
            service = await db.services.find_one({"id": service_id}, {"_id": 0})
            if service:
                vehicle_info = service.get('vehicle_number', 'N/A')
                service_type = service.get('service_type', 'Service')
                
                await create_activity(ActivityCreate(
                    type=ActivityType.SERVICE_COMPLETED,
                    title="Service completed",
                    description=f"{service_type} for {vehicle_info}",
                    icon="success",
                    metadata={"service_id": service_id}
                ))
        except Exception as e:
            logger.warning(f"Failed to create activity for service completion: {e}")
    
    return {"message": "Service status updated successfully"}

@api_router.get("/services/job-card/{job_card_number}")
async def get_service_by_job_card(job_card_number: str, current_user: User = Depends(get_current_user)):
    """Get service details by job card number for billing"""
    service = await db.services.find_one({"job_card_number": job_card_number.upper()})
    if not service:
        raise HTTPException(status_code=404, detail="Service not found with this job card number")
    
    # Get customer details — don't hard-fail if customer record missing
    customer = await db.customers.find_one({"id": service["customer_id"]}) or {}

    # Also try to get vehicle brand/model from vehicle record
    vehicle_brand = service.get("vehicle_brand", "")
    vehicle_model = service.get("vehicle_model", "")
    if service.get("vehicle_id") and (not vehicle_brand or not vehicle_model):
        vehicle = await db.vehicles.find_one({"id": service["vehicle_id"]}) or {}
        vehicle_brand = vehicle_brand or vehicle.get("brand", "")
        vehicle_model = vehicle_model or vehicle.get("model", "")

    service_details = {
        "service_id": service["id"],
        "job_card_number": service["job_card_number"],
        "customer_id": service["customer_id"],
        "customer_name": customer.get("name") or service.get("customer_name") or "",
        "customer_phone": customer.get("mobile") or customer.get("phone") or "",
        "customer_address": customer.get("address") or "",
        "vehicle_number": service["vehicle_number"],
        "vehicle_brand": vehicle_brand,
        "vehicle_model": vehicle_model,
        "vehicle_year": service.get("vehicle_year", ""),
        "service_type": service["service_type"],
        "description": service["description"],
        "service_date": service["service_date"],
        "amount": service["amount"],
        "status": service["status"],
        "created_at": service["created_at"]
    }
    
    return service_details

@api_router.delete("/services/{service_id}")
async def delete_service(service_id: str, current_user: User = Depends(get_current_user)):
    # Check if service exists
    existing_service = await db.services.find_one({"id": service_id})
    if not existing_service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    # Delete the service
    result = await db.services.delete_one({"id": service_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Service not found")
    
    return {"message": "Service deleted successfully", "deleted_service_id": service_id}

# Spare Parts endpoints
@api_router.post("/spare-parts", response_model=SparePart)
async def create_spare_part(spare_part_data: SparePartCreate, current_user: User = Depends(get_current_user)):
    spare_part = SparePart(**spare_part_data.dict())
    await db.spare_parts.insert_one(spare_part.dict())
    return spare_part

@api_router.get("/spare-parts", response_model=List[SparePart])
async def get_spare_parts(low_stock: bool = False, current_user: User = Depends(get_current_user)):
    filter_dict = {}
    if low_stock:
        filter_dict = {"$expr": {"$lte": ["$quantity", "$low_stock_threshold"]}}
    
    spare_parts = await db.spare_parts.find(filter_dict).to_list(1000)
    # Handle legacy spare parts that don't have GST fields
    processed_parts = []
    for part in spare_parts:
        # Add default values for missing GST fields
        if 'hsn_sac' not in part:
            part['hsn_sac'] = None
        if 'gst_percentage' not in part:
            part['gst_percentage'] = 18.0
        if 'unit' not in part:
            part['unit'] = 'Nos'
        if 'compatible_models' not in part:
            part['compatible_models'] = None
        processed_parts.append(SparePart(**part))
    return processed_parts

@api_router.post("/spare-parts/bills", response_model=SparePartBill)
async def create_spare_part_bill(bill_data: SparePartBillCreate, current_user: User = Depends(get_current_user)):
    # Generate bill number (atomic)
    seq = await next_sequence("spare_part_bills")
    bill_number = f"SPB-{seq:06d}"
    
    bill_dict = bill_data.dict()
    bill_dict['bill_number'] = bill_number
    bill_dict['created_by'] = current_user.id
    
    # Handle customer data - prioritize customer_data over customer_id
    if bill_dict.get('customer_data'):
        # Use the new customer data format
        bill_dict['customer_id'] = None  # Clear legacy field
    elif bill_dict.get('customer_id'):
        # For backwards compatibility, keep customer_id if no customer_data
        pass
    else:
        raise HTTPException(status_code=400, detail="Customer information is required")
    
    # Ensure all GST fields are present with defaults if not provided
    if 'subtotal' not in bill_dict:
        bill_dict['subtotal'] = 0
    if 'total_discount' not in bill_dict:
        bill_dict['total_discount'] = 0
    if 'total_cgst' not in bill_dict:
        bill_dict['total_cgst'] = 0
    if 'total_sgst' not in bill_dict:
        bill_dict['total_sgst'] = 0
    if 'total_tax' not in bill_dict:
        bill_dict['total_tax'] = 0
    if 'total_amount' not in bill_dict:
        bill_dict['total_amount'] = bill_dict['subtotal'] + bill_dict['total_tax'] - bill_dict['total_discount']
    
    bill = SparePartBill(**bill_dict)
    
    await db.spare_part_bills.insert_one(bill.dict())
    return bill

@api_router.get("/spare-parts/bills", response_model=List[SparePartBill])
async def get_spare_part_bills(current_user: User = Depends(get_current_user)):
    bills = await db.spare_part_bills.find().to_list(1000)
    # Handle legacy bills that don't have GST fields or customer data
    processed_bills = []
    for bill in bills:
        # Add default values for missing GST fields
        if 'subtotal' not in bill:
            bill['subtotal'] = bill.get('total_amount', 0)
        if 'total_discount' not in bill:
            bill['total_discount'] = 0
        if 'total_cgst' not in bill:
            bill['total_cgst'] = 0
        if 'total_sgst' not in bill:
            bill['total_sgst'] = 0
        if 'total_tax' not in bill:
            bill['total_tax'] = 0
        # Handle customer data backwards compatibility
        if 'customer_data' not in bill:
            bill['customer_data'] = None
        if 'customer_id' not in bill:
            bill['customer_id'] = None
        processed_bills.append(SparePartBill(**bill))
    return processed_bills

@api_router.delete("/spare-parts/bills/{bill_id}")
async def delete_spare_part_bill(bill_id: str, current_user: User = Depends(get_current_user)):
    # Check if spare part bill exists
    existing_bill = await db.spare_part_bills.find_one({"id": bill_id})
    if not existing_bill:
        raise HTTPException(status_code=404, detail="Spare part bill not found")
    
    # Delete the spare part bill
    result = await db.spare_part_bills.delete_one({"id": bill_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Spare part bill not found")
    
    return {"message": "Spare part bill deleted successfully", "deleted_bill_id": bill_id}

# Service Bills API Endpoints
@api_router.get("/service-bills/next-number")
async def get_next_service_bill_number(current_user: User = Depends(get_current_user)):
    """Return the next SB-XXXXXX number based on actual saved bills count (not a counter)."""
    # Count actual bills in DB — so deleting a bill doesn't skip numbers
    count = await db.service_bills.count_documents({})
    # Also peek at the highest existing SB number to avoid collisions
    last = await db.service_bills.find_one({}, sort=[("bill_number", -1)])
    last_seq = 0
    if last and last.get("bill_number", "").startswith("SB-"):
        try:
            last_seq = int(last["bill_number"][3:])
        except ValueError:
            pass
    next_seq = max(count, last_seq) + 1
    return {"bill_number": f"SB-{next_seq:06d}"}

@api_router.post("/service-bills", response_model=ServiceBill)
async def create_service_bill(bill_data: ServiceBillCreate, current_user: User = Depends(get_current_user)):
    # Generate bill number if not provided
    if not bill_data.bill_number:
        seq = await next_sequence("service_bills")
        bill_data.bill_number = f"SB-{seq:06d}"
    
    # Get customer info if customer_id is provided
    customer_name = bill_data.customer_name
    customer_mobile = bill_data.customer_mobile
    
    if bill_data.customer_id and not customer_name:
        customer = await db.customers.find_one({"id": bill_data.customer_id})
        if customer:
            customer_name = customer.get("name", "")
            customer_mobile = customer.get("mobile", customer.get("phone", ""))
    
    # Parse bill date
    bill_date = datetime.now(timezone.utc)
    if bill_data.bill_date:
        try:
            bill_date = datetime.fromisoformat(bill_data.bill_date.replace('Z', '+00:00'))
        except (ValueError, TypeError):
            pass
    
    # Reduce spare part quantities for items that have spare_part_id
    spare_part_updates = []
    if bill_data.items:
        for item in bill_data.items:
            if isinstance(item, dict) and item.get("spare_part_id"):
                spare_part_id = item["spare_part_id"]
                qty_used = item.get("qty", 1)
                
                # Check if spare part exists and has enough quantity
                spare_part = await db.spare_parts.find_one({"id": spare_part_id})
                if spare_part:
                    current_qty = spare_part.get("quantity", 0)
                    new_qty = max(0, current_qty - qty_used)  # Don't go below 0
                    
                    # Update spare part quantity
                    await db.spare_parts.update_one(
                        {"id": spare_part_id},
                        {"$set": {"quantity": new_qty}}
                    )
                    spare_part_updates.append({
                        "part_id": spare_part_id,
                        "part_name": spare_part.get("name", "Unknown"),
                        "qty_used": qty_used,
                        "old_qty": current_qty,
                        "new_qty": new_qty
                    })
    
    bill = ServiceBill(
        bill_number=bill_data.bill_number,
        job_card_number=bill_data.job_card_number,
        customer_id=bill_data.customer_id,
        customer_name=customer_name,
        customer_mobile=customer_mobile,
        vehicle_number=bill_data.vehicle_number,
        vehicle_brand=bill_data.vehicle_brand,
        vehicle_model=bill_data.vehicle_model,
        items=bill_data.items,
        subtotal=bill_data.subtotal,
        total_discount=bill_data.total_discount,
        total_cgst=bill_data.total_cgst,
        total_sgst=bill_data.total_sgst,
        total_tax=bill_data.total_tax,
        total_amount=bill_data.total_amount,
        bill_date=bill_date,
        status=bill_data.status,
        created_by=current_user.username
    )
    
    await db.service_bills.insert_one(bill.dict())
    
    # Log spare part inventory updates
    if spare_part_updates:
        print(f"Spare part inventory updated for bill {bill.bill_number}: {spare_part_updates}")
    
    return bill

@api_router.get("/service-bills")
async def get_service_bills(current_user: User = Depends(get_current_user)):
    bills = await db.service_bills.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return bills

@api_router.get("/service-bills/{bill_id}")
async def get_service_bill(bill_id: str, current_user: User = Depends(get_current_user)):
    bill = await db.service_bills.find_one({"id": bill_id}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Service bill not found")
    return bill

@api_router.put("/service-bills/{bill_id}/status")
async def update_service_bill_status(bill_id: str, status_update: dict, current_user: User = Depends(get_current_user)):
    existing_bill = await db.service_bills.find_one({"id": bill_id})
    if not existing_bill:
        raise HTTPException(status_code=404, detail="Service bill not found")
    
    new_status = status_update.get("status", "unpaid")
    if new_status not in ["paid", "unpaid"]:
        raise HTTPException(status_code=400, detail="Invalid status. Must be 'paid' or 'unpaid'")
    
    await db.service_bills.update_one(
        {"id": bill_id},
        {"$set": {"status": new_status}}
    )
    
    return {"message": f"Bill status updated to {new_status}", "bill_id": bill_id, "status": new_status}

@api_router.put("/service-bills/{bill_id}")
async def update_service_bill(bill_id: str, bill_update: dict, current_user: User = Depends(get_current_user)):
    """Full update of service bill including items"""
    existing_bill = await db.service_bills.find_one({"id": bill_id})
    if not existing_bill:
        raise HTTPException(status_code=404, detail="Service bill not found")
    
    # Build update data
    update_data = {}
    
    if "bill_number" in bill_update:
        update_data["bill_number"] = bill_update["bill_number"]
        update_data["job_card_number"] = bill_update["bill_number"]  # Keep both in sync
    
    if "customer_name" in bill_update:
        update_data["customer_name"] = bill_update["customer_name"]
    
    if "vehicle_reg_no" in bill_update:
        update_data["vehicle_reg_no"] = bill_update["vehicle_reg_no"]
    
    if "status" in bill_update:
        update_data["status"] = bill_update["status"]
    
    if "amount" in bill_update:
        update_data["amount"] = bill_update["amount"]
    
    if "items" in bill_update:
        update_data["items"] = bill_update["items"]
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.service_bills.update_one(
        {"id": bill_id},
        {"$set": update_data}
    )
    
    # Return updated bill
    updated_bill = await db.service_bills.find_one({"id": bill_id}, {"_id": 0})
    return updated_bill

@api_router.delete("/service-bills/{bill_id}")
async def delete_service_bill(bill_id: str, current_user: User = Depends(get_current_user)):
    existing_bill = await db.service_bills.find_one({"id": bill_id})
    if not existing_bill:
        raise HTTPException(status_code=404, detail="Service bill not found")
    
    result = await db.service_bills.delete_one({"id": bill_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Service bill not found")
    
    return {"message": "Service bill deleted successfully", "deleted_bill_id": bill_id}



# ─── Sales Milestone Tracker ──────────────────────────────────────────────────

@api_router.get("/sale-milestones")
async def list_sale_milestones(current_user: User = Depends(get_current_user)):
    """List all sale milestones (summary only, no image data)."""
    docs = await db.sale_milestones.find({}, {"_id": 0}).sort("updated_at", -1).to_list(1000)
    # Strip image data from listing to keep response small
    result = []
    for doc in docs:
        summary = {k: v for k, v in doc.items() if k != "milestones"}
        milestone_summary = {}
        for ms_key, ms_val in doc.get("milestones", {}).items():
            milestone_summary[ms_key] = {
                "completed": ms_val.get("completed", False),
                "completed_at": ms_val.get("completed_at"),
                "doc_count": len(ms_val.get("docs", [])),
                "notes": ms_val.get("notes", ""),
            }
        summary["milestones"] = milestone_summary
        result.append(summary)
    return result

@api_router.get("/sale-milestones/{sale_id}")
async def get_sale_milestone(sale_id: str, current_user: User = Depends(get_current_user)):
    """Get full milestone record for a sale including image data."""
    doc = await db.sale_milestones.find_one({"sale_id": sale_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Milestone not found")
    return doc

@api_router.post("/sale-milestones/{sale_id}/milestone/{milestone_key}/complete")
async def complete_milestone(
    sale_id: str,
    milestone_key: str,
    body: Dict[str, Any],
    current_user: User = Depends(get_current_user)
):
    """Mark a milestone complete with optional notes."""
    valid_keys = {"customer_docs", "invoice_insurance", "road_tax", "number_plates", "plates_attached"}
    if milestone_key not in valid_keys:
        raise HTTPException(status_code=400, detail=f"Invalid milestone key. Must be one of: {valid_keys}")

    now = datetime.now(timezone.utc)
    update = {
        f"milestones.{milestone_key}.completed": True,
        f"milestones.{milestone_key}.completed_at": now,
        f"milestones.{milestone_key}.notes": body.get("notes", ""),
        "updated_at": now,
    }

    result = await db.sale_milestones.update_one(
        {"sale_id": sale_id},
        {"$set": update},
        upsert=True
    )
    if result.upserted_id:
        sale = await db.sales.find_one({"id": sale_id}) or {}
        customer = await db.customers.find_one({"id": sale.get("customer_id", "")}) or {}
        await db.sale_milestones.update_one(
            {"sale_id": sale_id},
            {"$set": {
                "id": str(__import__("uuid").uuid4()),
                "sale_id": sale_id,
                "customer_name": customer.get("name") or sale.get("customer_name", ""),
                "invoice_number": sale.get("invoice_number", ""),
                "created_at": now,
            }},
            upsert=True
        )
    return {"message": "Milestone updated", "milestone_key": milestone_key}

@api_router.post("/sale-milestones/{sale_id}/milestone/{milestone_key}/uncomplete")
async def uncomplete_milestone(
    sale_id: str,
    milestone_key: str,
    current_user: User = Depends(get_current_user)
):
    """Unmark a milestone."""
    now = datetime.now(timezone.utc)
    await db.sale_milestones.update_one(
        {"sale_id": sale_id},
        {"$set": {
            f"milestones.{milestone_key}.completed": False,
            f"milestones.{milestone_key}.completed_at": None,
            "updated_at": now,
        }},
        upsert=True
    )
    return {"message": "Milestone unmarked"}

@api_router.post("/sale-milestones/init/{sale_id}")
async def init_sale_milestone(sale_id: str, current_user: User = Depends(get_current_user)):
    """Create an empty milestone record for a sale."""
    existing = await db.sale_milestones.find_one({"sale_id": sale_id})
    if existing:
        return {"message": "Already exists"}
    sale = await db.sales.find_one({"id": sale_id}) or {}
    customer = await db.customers.find_one({"id": sale.get("customer_id", "")}) or {}
    now = datetime.now(timezone.utc)
    rec = {
        "id": str(__import__("uuid").uuid4()),
        "sale_id": sale_id,
        "customer_name": customer.get("name", ""),
        "invoice_number": sale.get("invoice_number", ""),
        "milestones": {},
        "created_at": now,
        "updated_at": now,
    }
    await db.sale_milestones.insert_one(rec)
    return {"message": "Created", "id": rec["id"]}

@api_router.get("/service-report")
async def get_service_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Service report: revenue breakdown by parts vs labour, per job card"""
    query = {}
    if start_date and end_date:
        try:
            sd = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
            ed = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
            query["created_at"] = {"$gte": sd, "$lte": ed}
        except Exception:
            pass

    bills = await db.service_bills.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)

    report_rows = []
    total_parts_revenue = 0.0
    total_labour_revenue = 0.0
    total_tax = 0.0
    total_grand = 0.0

    for bill in bills:
        items = bill.get("items", [])
        parts_revenue = 0.0
        labour_revenue = 0.0
        bill_tax = 0.0

        for item in items:
            qty = float(item.get("qty", 1) or 1)
            rate = float(item.get("rate", 0) or 0)
            labor = float(item.get("labor", 0) or 0)
            disc = float(item.get("disc_percent", 0) or 0)
            gst = float(item.get("gst_percent", 0) or 0)

            parts_base = qty * rate
            labour_base = labor
            parts_after_disc = parts_base * (1 - disc / 100)
            labour_after_disc = labour_base * (1 - disc / 100)
            parts_tax = parts_after_disc * gst / 100
            labour_tax = labour_after_disc * gst / 100

            parts_revenue += parts_after_disc + parts_tax
            labour_revenue += labour_after_disc + labour_tax
            bill_tax += float(item.get("total_tax", 0) or 0)

        grand = float(bill.get("total_amount", 0) or bill.get("amount", 0) or 0)

        report_rows.append({
            "bill_id": bill.get("id"),
            "bill_number": bill.get("bill_number"),
            "job_card_number": bill.get("job_card_number"),
            "customer_name": bill.get("customer_name"),
            "customer_mobile": bill.get("customer_mobile"),
            "vehicle_number": bill.get("vehicle_number"),
            "bill_date": bill.get("bill_date") or bill.get("created_at"),
            "status": bill.get("status"),
            "parts_revenue": round(parts_revenue, 2),
            "labour_revenue": round(labour_revenue, 2),
            "total_tax": round(bill_tax, 2),
            "grand_total": round(grand, 2),
            "items_count": len(items),
        })

        total_parts_revenue += parts_revenue
        total_labour_revenue += labour_revenue
        total_tax += bill_tax
        total_grand += grand

    return {
        "rows": report_rows,
        "summary": {
            "total_bills": len(report_rows),
            "total_parts_revenue": round(total_parts_revenue, 2),
            "total_labour_revenue": round(total_labour_revenue, 2),
            "total_tax": round(total_tax, 2),
            "grand_total": round(total_grand, 2),
        }
    }


async def get_spare_part(part_id: str, current_user: User = Depends(get_current_user)):
    part = await db.spare_parts.find_one({"id": part_id})
    if not part:
        raise HTTPException(status_code=404, detail="Spare part not found")
    
    # Handle legacy spare parts that don't have GST fields
    if 'hsn_sac' not in part:
        part['hsn_sac'] = None
    if 'gst_percentage' not in part:
        part['gst_percentage'] = 18.0
    if 'unit' not in part:
        part['unit'] = 'Nos'
    if 'compatible_models' not in part:
        part['compatible_models'] = None
    
    return SparePart(**part)

@api_router.put("/spare-parts/{part_id}", response_model=SparePart)
async def update_spare_part(part_id: str, spare_part_data: SparePartCreate, current_user: User = Depends(get_current_user)):
    # Check if spare part exists
    existing_part = await db.spare_parts.find_one({"id": part_id})
    if not existing_part:
        raise HTTPException(status_code=404, detail="Spare part not found")
    
    # Update spare part data
    update_data = spare_part_data.dict()
    update_data["id"] = part_id  # Keep the original ID
    update_data["created_at"] = existing_part["created_at"]  # Keep original creation date
    
    updated_part = SparePart(**update_data)
    await db.spare_parts.replace_one({"id": part_id}, updated_part.dict())
    return updated_part

@api_router.delete("/spare-parts/{part_id}")
async def delete_spare_part(part_id: str, current_user: User = Depends(get_current_user)):
    # Check if spare part exists
    existing_part = await db.spare_parts.find_one({"id": part_id})
    if not existing_part:
        raise HTTPException(status_code=404, detail="Spare part not found")
    
    # Check if spare part is referenced in any bills
    bills_count = await db.spare_part_bills.count_documents({"items.part_id": part_id})
    if bills_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete spare part. Part is referenced in {bills_count} bill(s). Please remove from bills first.")
    
    # Delete the spare part
    result = await db.spare_parts.delete_one({"id": part_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Spare part not found")
    
    return {"message": "Spare part deleted successfully", "deleted_part_id": part_id}

# Dashboard stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    total_customers = await db.customers.count_documents({})
    total_vehicles = await db.vehicles.count_documents({})
    vehicles_in_stock = await db.vehicles.count_documents({"status": VehicleStatus.IN_STOCK})
    vehicles_sold = await db.vehicles.count_documents({"status": VehicleStatus.SOLD})
    pending_services = await db.services.count_documents({"status": ServiceStatus.PENDING})
    low_stock_parts = await db.spare_parts.count_documents({"$expr": {"$lte": ["$quantity", "$low_stock_threshold"]}})
    
    # Calculate completed services today
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59, microsecond=999999)
    completed_today = await db.services.count_documents({
        "status": ServiceStatus.COMPLETED,
        "updated_at": {"$gte": today_start, "$lte": today_end}
    })
    
    # Sales statistics including imported data
    total_sales = await db.sales.count_documents({})
    direct_sales = await db.sales.count_documents({"$or": [{"source": {"$exists": False}}, {"source": "direct"}]})
    imported_sales = await db.sales.count_documents({"source": "import"})
    
    # Calculate total sales revenue (including imported sales)
    sales_pipeline = [
        {"$group": {
            "_id": None,
            "total_revenue": {"$sum": "$amount"},
            "direct_revenue": {"$sum": {"$cond": [
                {"$or": [{"$not": ["$source"]}, {"$eq": ["$source", "direct"]}]},
                "$amount", 
                0
            ]}},
            "imported_revenue": {"$sum": {"$cond": [
                {"$eq": ["$source", "import"]},
                "$amount", 
                0
            ]}}
        }}
    ]
    
    revenue_stats = await db.sales.aggregate(sales_pipeline).to_list(1)
    total_revenue = revenue_stats[0]["total_revenue"] if revenue_stats else 0
    direct_revenue = revenue_stats[0]["direct_revenue"] if revenue_stats else 0
    imported_revenue = revenue_stats[0]["imported_revenue"] if revenue_stats else 0

    # Service bill statistics
    total_service_bills = await db.service_bills.count_documents({})
    paid_service_bills = await db.service_bills.count_documents({"status": "paid"})
    pending_service_bills = await db.service_bills.count_documents({"status": "pending"})
    service_bill_pipeline = [
        {"$group": {
            "_id": None,
            "total_service_revenue": {"$sum": "$total_amount"},
            "paid_revenue": {"$sum": {"$cond": [{"$eq": ["$status", "paid"]}, "$total_amount", 0]}},
            "pending_revenue": {"$sum": {"$cond": [{"$eq": ["$status", "pending"]}, "$total_amount", 0]}}
        }}
    ]
    service_bill_stats = await db.service_bills.aggregate(service_bill_pipeline).to_list(1)
    total_service_revenue = service_bill_stats[0]["total_service_revenue"] if service_bill_stats else 0
    paid_service_revenue = service_bill_stats[0]["paid_revenue"] if service_bill_stats else 0
    pending_service_revenue = service_bill_stats[0]["pending_revenue"] if service_bill_stats else 0

    # Total service jobs (job cards)
    total_service_jobs = await db.services.count_documents({})
    completed_services = await db.services.count_documents({"status": ServiceStatus.COMPLETED})
    in_progress_services = await db.services.count_documents({"status": ServiceStatus.IN_PROGRESS})

    return {
        "total_customers": total_customers,
        "total_vehicles": total_vehicles,
        "vehicles_in_stock": vehicles_in_stock,
        "vehicles_sold": vehicles_sold,
        "pending_services": pending_services,
        "completed_today": completed_today,
        "low_stock_parts": low_stock_parts,
        "sales_stats": {
            "total_sales": total_sales,
            "direct_sales": direct_sales,
            "imported_sales": imported_sales,
            "total_revenue": total_revenue,
            "direct_revenue": direct_revenue,
            "imported_revenue": imported_revenue
        },
        "service_stats": {
            "total_service_bills": total_service_bills,
            "paid_bills": paid_service_bills,
            "pending_bills": pending_service_bills,
            "total_service_revenue": total_service_revenue,
            "paid_revenue": paid_service_revenue,
            "pending_revenue": pending_service_revenue,
            "total_service_jobs": total_service_jobs,
            "completed_services": completed_services,
            "in_progress_services": in_progress_services,
            "pending_services": pending_services
        }
    }

# Import/Export endpoints
@api_router.post("/import/upload", response_model=ImportResult)
async def upload_import_file(
    data_type: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload and process import file"""
    
    # Validate data type
    valid_types = ["customers", "vehicles", "spare_parts", "services", "service_bills", "registrations", "service_customers"]
    if data_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid data type. Must be one of: {valid_types}")
    
    # Validate file format
    if not file.filename or not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    
    # Create import job
    import_job = ImportJob(
        file_name=file.filename,
        data_type=data_type,
        status="processing",
        created_by=current_user.id
    )
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Parse file based on type
        if file.filename.endswith('.csv'):
            data = await parse_csv_file(file_content)
        else:
            data = await parse_excel_file(file_content)
        
        import_job.total_records = len(data)
        
        # Process import based on data type
        if data_type == "customers":
            result = await import_customers_data(data, import_job, current_user.id)
        elif data_type == "vehicles":
            result = await import_vehicles_data(data, import_job, current_user.id)
        elif data_type == "spare_parts":
            result = await import_spare_parts_data(data, import_job, current_user.id)
        elif data_type == "services":
            result = await import_services_data(data, import_job, current_user.id)
        elif data_type == "service_bills":
            result = await import_service_bills_data(data, import_job, current_user.id)
        elif data_type == "registrations":
            result = await import_registrations_data(data, import_job, current_user.id)
        elif data_type == "service_customers":
            result = await import_service_customers_data(data, import_job, current_user.id)
        
        import_job.status = "completed"
        import_job.end_time = datetime.now(timezone.utc)
        
    except Exception as e:
        import_job.status = "failed"
        import_job.end_time = datetime.now(timezone.utc)
        import_job.errors.append({"error": str(e), "row": 0})
        result = ImportResult(
            job_id=import_job.id,
            status="failed",
            message=str(e)
        )
    
    # Save import job to database
    await db.import_jobs.insert_one(import_job.dict())
    
    return result

@api_router.get("/import/jobs", response_model=List[ImportJob])
async def get_import_jobs(
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get import job history"""
    jobs = await db.import_jobs.find().skip(skip).limit(limit).sort("created_at", -1).to_list(length=None)
    return [ImportJob(**job) for job in jobs]

@api_router.get("/import/template/{data_type}")
async def download_import_template(
    data_type: str,
    current_user: User = Depends(get_current_user)
):
    """Download CSV template for import"""
    from fastapi.responses import Response
    
    templates = {
        "customers": "name,care_of,mobile,email,address,brand,model,color,vehicle_number,chassis_number,engine_number,nominee_name,relation,age,sale_amount,payment_method,hypothecation,sale_date,invoice_number\nJohn Doe,S/O Ramesh,9876543210,john@example.com,\"123 Main St, Bangalore\",TVS,Apache RTR 160,Red,KA01AB1234,ABC123456789012345,ENG987654321,Jane Doe,spouse,28,75000,cash,cash,2024-01-15,INV001\nJane Smith,D/O Kumar,9876543211,jane@example.com,\"456 Oak Ave, Mysore\",BAJAJ,Pulsar 150,Blue,KA02CD5678,DEF123456789012345,ENG987654322,John Smith,father,55,65000,finance,\"Bank Finance\",2024-01-16,INV002",
        "vehicles": "date_received,brand,model,chassis_number,engine_number,color,vehicle_number,key_number,inbound_location,page_number,status,customer_mobile,customer_name,sale_amount,payment_method\n2025-01-15,TVS,Apache RTR 160,ABC123456789,ENG987654321,Red,KA01AB1234,KEY001,Warehouse A,Page 1,in_stock,9876543210,John Doe,75000,cash\n2025-01-16,BAJAJ,Pulsar 150,DEF123456789,ENG987654322,Blue,KA02CD5678,KEY002,Warehouse B,Page 2,in_stock,9876543211,Jane Smith,65000,finance",
        "spare_parts": "name,part_number,brand,quantity,unit,unit_price,hsn_sac,gst_percentage,supplier,compatible_models\nBrake Pad,BP001,TVS,50,Nos,250.00,87084090,18.0,ABC Supplies,\"Apache RTR 160, Pulsar 150\"\nEngine Oil,EO001,CASTROL,25,Ltr,450.00,27101981,28.0,XYZ Motors,\"All Models\"",
        "services": "registration_date,customer_name,customer_mobile,vehicle_number,chassis_number,vehicle_brand,vehicle_model,vehicle_year,service_type,description,amount\n2025-01-15,John Doe,9876543210,KA01AB1234,ABC123456789,TVS,Apache RTR 160,2024,periodic_service,General servicing,1500.00\n2025-01-16,Jane Smith,9876543211,KA02CD5678,DEF123456789,BAJAJ,Pulsar 150,2023,repair,Brake repair,800.00",
        "service_bills": "bill_date,customer_name,customer_mobile,vehicle_number,vehicle_brand,vehicle_model,job_card_number,item_description,item_hsn,item_qty,item_rate,item_gst_percent,total_amount,status\n2025-01-15,John Doe,9876543210,KA01AB1234,TVS,Apache RTR 160,JOB-000001,Engine Oil Change,27101,1,450,18,531,paid\n2025-01-16,Jane Smith,9876543211,KA02CD5678,BAJAJ,Pulsar 150,,Brake Pad Replacement,87084090,2,250,18,590,pending",
        "registrations": "customer_name,customer_mobile,customer_address,vehicle_number,vehicle_brand,vehicle_model,vehicle_year,chassis_number,engine_number,registration_date\nJohn Doe,9876543210,\"123 Main St, Bangalore\",KA01AB1234,TVS,Apache RTR 160,2024,ABC123456789012345,ENG987654321,2025-01-15\nJane Smith,9876543211,\"456 Oak Ave, Mysore\",KA02CD5678,BAJAJ,Pulsar 150,2023,DEF123456789012345,ENG987654322,2025-01-16",
        "service_customers": "name,mobile,phone,email,address,vehicle_brand,vehicle_model,vehicle_year,vehicle_number,chassis_number,engine_number\nJohn Doe,9876543210,,john@example.com,\"123 Main St, Bangalore\",TVS,Apache RTR 160,2024,KA01AB1234,ABC123456789012345,ENG987654321\nJane Smith,9876543211,9876543212,jane@example.com,\"456 Oak Ave, Mysore\",BAJAJ,Pulsar 150,2023,KA02CD5678,DEF123456789012345,ENG987654322"
    }
    
    if data_type not in templates:
        raise HTTPException(status_code=404, detail="Template not found")
    
    return Response(
        content=templates[data_type],
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={data_type}_template.csv"}
    )

# Helper functions for file parsing and data import
async def parse_csv_file(file_content: bytes) -> List[Dict]:
    """Parse CSV file content with multiple encoding support"""
    # List of encodings to try in order of preference
    encodings = ['utf-8', 'iso-8859-1', 'windows-1252', 'cp1252']
    
    content = None
    successful_encoding = None
    
    for encoding in encodings:
        try:
            content = file_content.decode(encoding)
            successful_encoding = encoding
            break
        except UnicodeDecodeError:
            continue
    
    # If all encodings fail, use UTF-8 with error handling
    if content is None:
        content = file_content.decode('utf-8', errors='replace')
        successful_encoding = 'utf-8 (with error replacement)'
    
    # Log which encoding was used for debugging
    logging.info(f"CSV file parsed successfully using encoding: {successful_encoding}")
    
    csv_reader = csv.DictReader(io.StringIO(content))
    return list(csv_reader)

# Cross-reference utility functions for unified data import
async def find_customer_by_mobile(mobile: str) -> Optional[Dict]:
    """Find existing customer by mobile number"""
    if not mobile or mobile == "0000000000":
        return None
    return await db.customers.find_one({"mobile": mobile})

async def find_vehicle_by_identifiers(vehicle_number: Optional[str] = None, chassis_number: Optional[str] = None) -> Optional[Dict]:
    """Find existing vehicle by vehicle_number or chassis_number"""
    if not vehicle_number and not chassis_number:
        return None
    
    query = []
    if vehicle_number:
        query.append({"vehicle_number": vehicle_number})
    if chassis_number:
        query.append({"chassis_number": chassis_number})
    
    if query:
        return await db.vehicles.find_one({"$or": query})
    return None

async def find_or_create_customer(mobile: str, data: Dict, import_stats: Dict) -> str:
    """Find existing customer by mobile or create new one"""
    # Try to find existing customer
    existing_customer = await find_customer_by_mobile(mobile)
    if existing_customer:
        import_stats['customers_linked'] = import_stats.get('customers_linked', 0) + 1
        return existing_customer['id']
    
    # Create new customer with available data
    customer_data = CustomerCreate(
        name=data.get('name', 'Unknown Customer'),
        mobile=mobile,
        email=data.get('email') or None,
        address=data.get('address', 'Address not provided'),
        care_of=data.get('care_of') or None
    )
    
    customer = Customer(**customer_data.dict())
    await db.customers.insert_one(customer.dict())
    import_stats['customers_created'] = import_stats.get('customers_created', 0) + 1
    return customer.id

async def find_or_create_vehicle(vehicle_number: Optional[str], chassis_number: Optional[str], data: Dict, import_stats: Dict) -> Optional[str]:
    """Find existing vehicle or create new one"""
    # Try to find existing vehicle
    existing_vehicle = await find_vehicle_by_identifiers(vehicle_number, chassis_number)
    if existing_vehicle:
        import_stats['vehicles_linked'] = import_stats.get('vehicles_linked', 0) + 1
        return existing_vehicle['id']
    
    # Create new vehicle if we have enough data
    if not chassis_number and not vehicle_number:
        return None
    
    vehicle_data = VehicleCreate(
        brand=data.get('brand', 'UNKNOWN'),
        model=data.get('model', 'Unknown Model'),
        chassis_number=chassis_number or f'AUTO-{str(uuid.uuid4())[:8]}',
        engine_number=data.get('engine_number', 'Unknown Engine'),
        color=data.get('color', 'Unknown Color'),
        vehicle_number=vehicle_number,
        key_number=data.get('key_number')
    )
    
    vehicle = Vehicle(**vehicle_data.dict())
    await db.vehicles.insert_one(vehicle.dict())
    import_stats['vehicles_created'] = import_stats.get('vehicles_created', 0) + 1
    return vehicle.id

async def check_customer_duplicate(mobile: str) -> bool:
    """Check if customer with mobile number already exists"""
    existing = await find_customer_by_mobile(mobile)
    return existing is not None

async def check_vehicle_duplicate(chassis_number: str) -> bool:
    """Check if vehicle with chassis number already exists"""
    existing = await db.vehicles.find_one({"chassis_number": chassis_number})
    return existing is not None

async def parse_excel_file(file_content: bytes) -> List[Dict]:
    """Parse Excel file content"""
    df = pd.read_excel(io.BytesIO(file_content))
    return df.to_dict('records')

async def import_customers_data(data: List[Dict], import_job: ImportJob, user_id: str) -> ImportResult:
    """Import customers data with smart merge logic:
    - New customer (mobile not found): insert everything
    - Existing customer (mobile found):
        * Fill only missing fields (name, email, address, care_of, vehicle_info, insurance_info, sales_info)
        * If same vehicle model already on record → skip entirely
        * If different vehicle model → update with new vehicle details
    """
    successful = 0
    failed = 0
    skipped = 0
    updated = 0
    errors = []
    incomplete_records = []
    import_stats = {'vehicles_linked': 0, 'sales_created': 0, 'customers_updated': 0}

    # ── PRE-FETCH ──────────────────────────────────────────────────────────────
    # mobile -> full customer doc
    existing_customers = {}
    async for c in db.customers.find({}, {"_id": 0}):
        if c.get("mobile"):
            existing_customers[str(c["mobile"])] = c

    existing_vehicles = {}        # chassis_number -> vehicle doc
    existing_vehicles_by_reg = {} # vehicle_number -> vehicle doc
    async for v in db.vehicles.find({}, {"id": 1, "chassis_number": 1, "vehicle_number": 1}):
        if v.get("chassis_number"):
            existing_vehicles[v["chassis_number"]] = v
        if v.get("vehicle_number"):
            existing_vehicles_by_reg[v["vehicle_number"]] = v

    # Track mobiles seen in this file to catch within-file dupes
    seen_in_file = {}  # mobile -> vehicle_model (first occurrence wins per model)

    customers_to_insert = []
    sales_to_insert = []
    vehicle_updates = []
    customer_bulk_updates = []  # (filter, update) for existing customers

    for idx, row in enumerate(data):
        try:
            phone_number = safe_str(row.get('mobile', '')) or safe_str(row.get('phone', ''))
            name = safe_str(row.get('name', '')) or "Customer"
            if not phone_number:
                phone_number = "0000000000"
            address = safe_str(row.get('address', ''))

            # ── Build sub-dicts ────────────────────────────────────────────────
            vehicle_info = {}
            if any(row.get(k) for k in ['brand', 'model', 'vehicle_no', 'vehicle_number', 'chassis_no', 'chassis_number']):
                vehicle_info = {
                    'brand':          safe_str(row.get('brand', '')),
                    'model':          safe_str(row.get('model', '')),
                    'color':          safe_str(row.get('color', '')),
                    'vehicle_number': safe_str(row.get('vehicle_number', '')) or safe_str(row.get('vehicle_no', '')),
                    'chassis_number': safe_str(row.get('chassis_number', '')) or safe_str(row.get('chassis_no', '')),
                    'engine_number':  safe_str(row.get('engine_number', '')) or safe_str(row.get('engine_no', '')),
                }

            insurance_info = {}
            if any(row.get(k) for k in ['nominee_name', 'relation', 'age']):
                insurance_info = {
                    'nominee_name': safe_str(row.get('nominee_name', '')),
                    'relation':     safe_str(row.get('relation', '')),
                    'age':          safe_str(row.get('age', '')),
                }

            sales_info = {}
            if any(row.get(k) for k in ['sale_amount', 'payment_method']):
                sales_info = {
                    'amount':         safe_str(row.get('sale_amount', '')),
                    'payment_method': safe_str(row.get('payment_method', '')),
                    'hypothecation':  safe_str(row.get('hypothecation', '')),
                    'sale_date':      safe_str(row.get('sale_date', '')),
                    'invoice_number': safe_str(row.get('invoice_number', '')),
                }

            incoming_model = vehicle_info.get('model', '').strip().lower()
            chassis = vehicle_info.get('chassis_number', '')
            reg = vehicle_info.get('vehicle_number', '')

            # ── LOGIC 2: within-file duplicate check ───────────────────────────
            if phone_number != "0000000000":
                if phone_number in seen_in_file:
                    existing_models = seen_in_file[phone_number]
                    if incoming_model and incoming_model in existing_models:
                        # Same mobile + same model already seen in this file → skip
                        skipped += 1
                        continue
                    # Different model (or no model) → allow through, track it
                    if incoming_model:
                        seen_in_file[phone_number].add(incoming_model)
                else:
                    seen_in_file[phone_number] = {incoming_model} if incoming_model else set()

            # ── LOGIC 1 & 2: existing customer in DB ──────────────────────────
            if phone_number != "0000000000" and phone_number in existing_customers:
                existing = existing_customers[phone_number]
                existing_vehicle_info = existing.get('vehicle_info') or {}
                existing_model = existing_vehicle_info.get('model', '').strip().lower()

                # LOGIC 2: same vehicle model → skip entirely
                if incoming_model and existing_model and incoming_model == existing_model:
                    skipped += 1
                    continue

                # LOGIC 2: different vehicle model → create a NEW customer record
                # with the new vehicle details (same name/mobile but separate entry)
                if incoming_model and existing_model and incoming_model != existing_model:
                    new_customer_data = CustomerCreate(
                        name=name,
                        mobile=phone_number,
                        email=safe_str(row.get('email', '')) or None,
                        address=address or existing.get('address') or "Address not provided",
                        care_of=safe_str(row.get('care_of', '')) or existing.get('care_of') or None
                    )
                    new_customer = Customer(**new_customer_data.dict())
                    new_customer_dict = new_customer.dict()

                    if vehicle_info and any(vehicle_info.values()):
                        new_customer_dict['vehicle_info'] = vehicle_info
                    if insurance_info and any(insurance_info.values()):
                        new_customer_dict['insurance_info'] = insurance_info
                    elif existing.get('insurance_info'):
                        new_customer_dict['insurance_info'] = existing['insurance_info']
                    if sales_info and any(sales_info.values()):
                        new_customer_dict['sales_info'] = sales_info

                    customers_to_insert.append(new_customer_dict)

                    # Link vehicle for the new customer
                    matched_vehicle = existing_vehicles.get(chassis) or existing_vehicles_by_reg.get(reg)
                    if matched_vehicle:
                        vehicle_updates.append(
                            ({"id": matched_vehicle['id']}, {"$set": {"customer_id": new_customer.id}})
                        )
                        import_stats['vehicles_linked'] += 1

                    # Build sale for new customer if sales info present
                    if sales_info and sales_info.get('amount'):
                        try:
                            sale_date = None
                            date_str = str(sales_info.get('sale_date', '')).strip()
                            if date_str:
                                try:
                                    from datetime import datetime as _dt
                                    import re as _re
                                    if date_str.replace('.', '').isdigit():
                                        excel_date = float(date_str)
                                        if excel_date > 60: excel_date -= 1
                                        sale_date = _dt(1900, 1, 1) + timedelta(days=excel_date - 2)
                                    elif _re.match(r'\d{1,2}-[A-Za-z]{3}', date_str):
                                        sale_date = _dt.strptime(f"{date_str}-{_dt.now().year}", "%d-%b-%Y")
                                    elif _re.match(r'\d{1,2}/\d{1,2}/\d{4}', date_str):
                                        sale_date = _dt.strptime(date_str, "%d/%m/%Y")
                                    elif _re.match(r'\d{1,2}-\d{1,2}-\d{4}', date_str):
                                        sale_date = _dt.strptime(date_str, "%d-%m-%Y")
                                    elif _re.match(r'\d{4}-\d{1,2}-\d{1,2}', date_str):
                                        sale_date = _dt.strptime(date_str, "%Y-%m-%d")
                                    else:
                                        from dateutil import parser as _dp
                                        sale_date = _dp.parse(date_str)
                                except Exception:
                                    sale_date = datetime(datetime.now().year, 1, 1)
                            if not sale_date:
                                sale_date = datetime(datetime.now().year, 1, 1)
                            vehicle_id = existing_vehicles.get(chassis, {}).get('id') if chassis else None
                            sale_record = Sale(
                                customer_id=new_customer.id,
                                vehicle_id=vehicle_id,
                                amount=float(sales_info['amount']),
                                payment_method=(sales_info.get('payment_method') or 'CASH').upper(),
                                hypothecation=sales_info.get('hypothecation', ''),
                                sale_date=sale_date,
                                invoice_number=sales_info.get('invoice_number') or f"IMP-{new_customer.id[:8]}",
                                vehicle_brand=vehicle_info.get('brand', ''),
                                vehicle_model=vehicle_info.get('model', ''),
                                vehicle_color=vehicle_info.get('color', ''),
                                vehicle_chassis=vehicle_info.get('chassis_number', ''),
                                vehicle_engine=vehicle_info.get('engine_number', ''),
                                vehicle_registration=vehicle_info.get('vehicle_number', ''),
                                insurance_nominee=insurance_info.get('nominee_name', ''),
                                insurance_relation=insurance_info.get('relation', ''),
                                insurance_age=insurance_info.get('age', ''),
                                source="import",
                                created_by="import_system"
                            )
                            sales_to_insert.append(sale_record.dict())
                            import_stats['sales_created'] += 1
                        except Exception as sale_error:
                            print(f"Warning: Could not build sale for row {idx+2}: {sale_error}")

                    successful += 1
                    continue

                # LOGIC 1: same mobile, no vehicle model conflict → fill only missing fields
                update_fields = {}

                if name and name != "Customer" and not existing.get('name'):
                    update_fields['name'] = name
                if address and not existing.get('address'):
                    update_fields['address'] = address
                if safe_str(row.get('email', '')) and not existing.get('email'):
                    update_fields['email'] = safe_str(row.get('email', ''))
                if safe_str(row.get('care_of', '')) and not existing.get('care_of'):
                    update_fields['care_of'] = safe_str(row.get('care_of', ''))

                if vehicle_info and any(vehicle_info.values()):
                    existing_vi = existing.get('vehicle_info') or {}
                    merged_vi = dict(existing_vi)
                    for k, v in vehicle_info.items():
                        if v and not merged_vi.get(k):
                            merged_vi[k] = v
                    if merged_vi != existing_vi:
                        update_fields['vehicle_info'] = merged_vi

                if insurance_info and any(insurance_info.values()):
                    existing_ii = existing.get('insurance_info') or {}
                    merged_ii = dict(existing_ii)
                    for k, v in insurance_info.items():
                        if v and not merged_ii.get(k):
                            merged_ii[k] = v
                    if merged_ii != existing_ii:
                        update_fields['insurance_info'] = merged_ii

                if sales_info and any(sales_info.values()):
                    existing_si = existing.get('sales_info') or {}
                    merged_si = dict(existing_si)
                    for k, v in sales_info.items():
                        if v and not merged_si.get(k):
                            merged_si[k] = v
                    if merged_si != existing_si:
                        update_fields['sales_info'] = merged_si

                if update_fields:
                    customer_bulk_updates.append(
                        ({"id": existing['id']}, {"$set": update_fields})
                    )
                    import_stats['customers_updated'] += 1
                    updated += 1
                    existing_customers[phone_number] = {**existing, **update_fields}
                else:
                    skipped += 1

                matched_vehicle = existing_vehicles.get(chassis) or existing_vehicles_by_reg.get(reg)
                if matched_vehicle and not matched_vehicle.get('customer_id'):
                    vehicle_updates.append(
                        ({"id": matched_vehicle['id']}, {"$set": {"customer_id": existing['id']}})
                    )
                    import_stats['vehicles_linked'] += 1

                successful += 1
                continue

            # ── NEW CUSTOMER ───────────────────────────────────────────────────
            customer_data = CustomerCreate(
                name=name, mobile=phone_number,
                email=safe_str(row.get('email', '')) or None,
                address=address or "Address not provided",
                care_of=safe_str(row.get('care_of', '')) or None
            )
            customer = Customer(**customer_data.dict())
            customer_dict = customer.dict()

            if vehicle_info and any(vehicle_info.values()):
                customer_dict['vehicle_info'] = vehicle_info
            if insurance_info and any(insurance_info.values()):
                customer_dict['insurance_info'] = insurance_info
            if sales_info and any(sales_info.values()):
                customer_dict['sales_info'] = sales_info

            customers_to_insert.append(customer_dict)
            # Cache so later rows in this file see it
            existing_customers[phone_number] = customer_dict

            # Queue vehicle link
            matched_vehicle = existing_vehicles.get(chassis) or existing_vehicles_by_reg.get(reg)
            if matched_vehicle:
                vehicle_updates.append(
                    ({"id": matched_vehicle['id']}, {"$set": {"customer_id": customer.id}})
                )
                import_stats['vehicles_linked'] += 1

            # Build sale record
            if sales_info and sales_info.get('amount'):
                try:
                    sale_date = None
                    date_str = str(sales_info.get('sale_date', '')).strip()
                    if date_str:
                        try:
                            from datetime import datetime as _dt
                            import re as _re
                            if date_str.replace('.', '').isdigit():
                                excel_date = float(date_str)
                                if excel_date > 60:
                                    excel_date -= 1
                                sale_date = _dt(1900, 1, 1) + timedelta(days=excel_date - 2)
                            elif _re.match(r'\d{1,2}-[A-Za-z]{3}', date_str):
                                sale_date = _dt.strptime(f"{date_str}-{_dt.now().year}", "%d-%b-%Y")
                            elif _re.match(r'\d{1,2}/\d{1,2}/\d{4}', date_str):
                                sale_date = _dt.strptime(date_str, "%d/%m/%Y")
                            elif _re.match(r'\d{1,2}-\d{1,2}-\d{4}', date_str):
                                sale_date = _dt.strptime(date_str, "%d-%m-%Y")
                            elif _re.match(r'\d{4}-\d{1,2}-\d{1,2}', date_str):
                                sale_date = _dt.strptime(date_str, "%Y-%m-%d")
                            else:
                                from dateutil import parser as _dp
                                sale_date = _dp.parse(date_str)
                        except Exception:
                            sale_date = datetime(datetime.now().year, 1, 1)
                    if not sale_date:
                        sale_date = datetime(datetime.now().year, 1, 1)

                    vehicle_id = None
                    if chassis and chassis in existing_vehicles:
                        vehicle_id = existing_vehicles[chassis]['id']

                    sale_record = Sale(
                        customer_id=customer.id,
                        vehicle_id=vehicle_id,
                        amount=float(sales_info['amount']),
                        payment_method=(sales_info.get('payment_method') or 'CASH').upper(),
                        hypothecation=sales_info.get('hypothecation', ''),
                        sale_date=sale_date,
                        invoice_number=sales_info.get('invoice_number') or f"IMP-{customer.id[:8]}",
                        vehicle_brand=vehicle_info.get('brand', ''),
                        vehicle_model=vehicle_info.get('model', ''),
                        vehicle_color=vehicle_info.get('color', ''),
                        vehicle_chassis=vehicle_info.get('chassis_number', ''),
                        vehicle_engine=vehicle_info.get('engine_number', ''),
                        vehicle_registration=vehicle_info.get('vehicle_number', ''),
                        insurance_nominee=insurance_info.get('nominee_name', ''),
                        insurance_relation=insurance_info.get('relation', ''),
                        insurance_age=insurance_info.get('age', ''),
                        source="import",
                        created_by="import_system"
                    )
                    sales_to_insert.append(sale_record.dict())
                    import_stats['sales_created'] += 1
                except Exception as sale_error:
                    print(f"Warning: Could not build sale for row {idx+2}: {sale_error}")

            missing_fields = []
            if not vehicle_info or not any(vehicle_info.values()):
                missing_fields.append('vehicle_info')
            if not insurance_info or not any(insurance_info.values()):
                missing_fields.append('insurance_info')
            if not sales_info or not any(sales_info.values()):
                missing_fields.append('sales_info')
            if missing_fields:
                incomplete_records.append({"record_id": customer.id, "row": idx + 2, "missing_fields": missing_fields, "data": row})

            successful += 1

        except Exception as e:
            failed += 1
            errors.append({"row": idx + 2, "data": row, "error": str(e)})

    # ── BULK WRITE ─────────────────────────────────────────────────────────────
    if customers_to_insert:
        await db.customers.insert_many(customers_to_insert, ordered=False)
    if sales_to_insert:
        await db.sales.insert_many(sales_to_insert, ordered=False)
    import asyncio as _asyncio
    if vehicle_updates:
        await _asyncio.gather(*[db.vehicles.update_one(f, u) for f, u in vehicle_updates])
    if customer_bulk_updates:
        await _asyncio.gather(*[db.customers.update_one(f, u) for f, u in customer_bulk_updates])

    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors
    import_job.cross_reference_stats = import_stats
    import_job.incomplete_records = incomplete_records

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Import completed: {successful} processed ({import_stats['customers_updated']} existing updated, {successful - import_stats['customers_updated']} new), {failed} failed, {skipped} skipped (same vehicle model). Vehicles linked: {import_stats['vehicles_linked']}, sales created: {import_stats['sales_created']}.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )

async def import_vehicles_data(data: List[Dict], import_job: ImportJob, user_id: str) -> ImportResult:
    """Import vehicles with smart merge logic:
    - New vehicle (chassis not found): insert everything
    - Existing vehicle (chassis found):
        * All fields same → skip
        * Missing fields found → fill only the missing ones
    """
    successful = 0
    failed = 0
    skipped = 0
    updated = 0
    errors = []
    incomplete_records = []
    import_stats = {'customers_linked': 0, 'customers_created': 0, 'sales_created': 0, 'vehicles_updated': 0}

    valid_brands = ["TVS", "BAJAJ", "HERO", "HONDA", "TRIUMPH", "KTM", "SUZUKI", "APRILIA", "YAMAHA", "PIAGGIO", "ROYAL ENFIELD"]

    # Pre-fetch all existing vehicles by chassis for O(1) lookup
    existing_vehicles = {}  # chassis_number -> full vehicle doc
    async for v in db.vehicles.find({}, {"_id": 0}):
        if v.get("chassis_number"):
            existing_vehicles[v["chassis_number"]] = v

    for idx, row in enumerate(data):
        try:
            brand = safe_str(row.get('brand', '')).upper() or 'UNKNOWN'
            if brand != 'UNKNOWN' and brand not in valid_brands:
                brand = 'UNKNOWN'

            chassis_number = safe_str(row.get('chassis_number', '')) or safe_str(row.get('chassis_no', ''))
            engine_number  = safe_str(row.get('engine_number', ''))  or safe_str(row.get('engine_no', ''))
            key_number     = safe_str(row.get('key_number', ''))     or safe_str(row.get('key_no', ''))
            vehicle_number = safe_str(row.get('vehicle_number', ''))
            model          = safe_str(row.get('model', '')) or 'Unknown Model'
            color          = safe_str(row.get('color', '')) or 'Unknown Color'
            inbound_loc    = safe_str(row.get('inbound_location', '')) or 'Unknown Location'
            page_number    = safe_str(row.get('page_number', '')) or None
            customer_mobile = safe_str(row.get('customer_mobile', ''))
            customer_name   = safe_str(row.get('customer_name', ''))
            sale_amount     = safe_str(row.get('sale_amount', ''))
            payment_method  = safe_str(row.get('payment_method', ''))

            status = safe_str(row.get('status', '')).lower()
            if status not in ['available', 'in_stock', 'sold', 'returned']:
                status = 'available'

            date_received = datetime.now(timezone.utc)
            date_str = safe_str(row.get('date_received', ''))
            if date_str:
                try:
                    date_received = parse_date_flexible(date_str)
                except Exception:
                    pass

            # ── EXISTING VEHICLE ───────────────────────────────────────────────
            if chassis_number and chassis_number in existing_vehicles:
                existing = existing_vehicles[chassis_number]

                # Build incoming field set (non-empty values only)
                incoming = {
                    'brand':           brand if brand != 'UNKNOWN' else None,
                    'model':           model if model != 'Unknown Model' else None,
                    'color':           color if color != 'Unknown Color' else None,
                    'engine_number':   engine_number or None,
                    'vehicle_number':  vehicle_number or None,
                    'key_number':      key_number or None,
                    'inbound_location': inbound_loc if inbound_loc != 'Unknown Location' else None,
                    'page_number':     page_number,
                }

                # Fill only fields that are blank/missing on existing doc
                update_fields = {}
                for field, val in incoming.items():
                    if val and not existing.get(field):
                        update_fields[field] = val

                if not update_fields:
                    # All fields same (or nothing new to add) → skip
                    skipped += 1
                    continue

                await db.vehicles.update_one({"id": existing['id']}, {"$set": update_fields})
                existing_vehicles[chassis_number] = {**existing, **update_fields}
                import_stats['vehicles_updated'] += 1
                updated += 1
                successful += 1
                continue

            # ── NEW VEHICLE ────────────────────────────────────────────────────
            vehicle_dict = {
                'id': str(uuid.uuid4()),
                'brand': brand,
                'model': model,
                'chassis_number': chassis_number or 'Unknown Chassis',
                'engine_number': engine_number or 'Unknown Engine',
                'color': color,
                'vehicle_number': vehicle_number or None,
                'key_number': key_number or 'Unknown Key',
                'inbound_location': inbound_loc,
                'page_number': page_number,
                'status': status,
                'date_received': date_received,
                'created_at': datetime.now(timezone.utc),
            }

            customer_id = None
            if customer_mobile:
                customer_id = await find_or_create_customer(
                    customer_mobile, {'name': customer_name or 'Unknown Customer'}, import_stats
                )
                vehicle_dict['customer_id'] = customer_id

            vehicle = Vehicle(**vehicle_dict)
            await db.vehicles.insert_one(vehicle.dict())
            existing_vehicles[chassis_number] = vehicle.dict()

            if sale_amount and customer_id:
                try:
                    sale_record = Sale(
                        customer_id=customer_id,
                        vehicle_id=vehicle.id,
                        amount=float(sale_amount),
                        payment_method=payment_method.upper() or 'CASH',
                        sale_date=datetime.now(timezone.utc),
                        invoice_number=f"IMP-VEH-{vehicle.id[:8]}",
                        vehicle_brand=brand,
                        vehicle_model=model,
                        vehicle_color=color,
                        vehicle_chassis=chassis_number,
                        vehicle_engine=engine_number,
                        vehicle_registration=vehicle_number,
                        source="import",
                        created_by=user_id
                    )
                    await db.sales.insert_one(sale_record.dict())
                    import_stats['sales_created'] += 1
                    await db.vehicles.update_one({"id": vehicle.id}, {"$set": {"status": "sold"}})
                except Exception as sale_error:
                    print(f"Warning: Could not create sale for vehicle {vehicle.id}: {sale_error}")

            if not customer_id and customer_mobile:
                incomplete_records.append({"record_id": vehicle.id, "row": idx + 2, "missing_fields": ["customer_details"], "data": row})

            successful += 1

        except Exception as e:
            failed += 1
            errors.append({"row": idx + 2, "data": row, "error": str(e)})

    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors
    import_job.cross_reference_stats = import_stats
    import_job.incomplete_records = incomplete_records

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Vehicles import: {successful} processed ({successful - updated} new, {updated} updated with missing fields), {failed} failed, {skipped} skipped (no new data). Customers linked: {import_stats['customers_linked']}, created: {import_stats['customers_created']}, sales: {import_stats['sales_created']}.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )


    successful = 0
    failed = 0
    skipped = 0
    errors = []
    incomplete_records = []
    import_stats = {
        'customers_linked': 0,
        'customers_created': 0,
        'sales_created': 0
    }
    
    valid_brands = ["TVS", "BAJAJ", "HERO", "HONDA", "TRIUMPH", "KTM", "SUZUKI", "APRILIA", "YAMAHA", "PIAGGIO", "ROYAL ENFIELD"]
    
    for idx, row in enumerate(data):
        try:
            # Get fields with fallback values - use safe_str to handle float/NaN values
            brand = safe_str(row.get('brand', '')).upper() or 'UNKNOWN'
            if brand != 'UNKNOWN' and brand not in valid_brands:
                brand = 'UNKNOWN'
            
            # Support both old and new field names for backward compatibility
            chassis_number = (safe_str(row.get('chassis_number', '')) or 
                            safe_str(row.get('chassis_no', '')))
            engine_number = (safe_str(row.get('engine_number', '')) or 
                           safe_str(row.get('engine_no', '')))
            key_number = (safe_str(row.get('key_number', '')) or 
                        safe_str(row.get('key_no', '')))
            vehicle_number = safe_str(row.get('vehicle_number', ''))
            
            # Handle status field with validation
            status = safe_str(row.get('status', '')).lower()
            valid_statuses = ['available', 'in_stock', 'sold', 'returned']
            if status not in valid_statuses:
                status = 'available'
            
            # Check for duplicate chassis number before inserting
            if chassis_number and chassis_number != 'Unknown Chassis':
                existing_vehicle = await db.vehicles.find_one({"chassis_number": chassis_number})
                if existing_vehicle:
                    # Skip duplicate vehicle (don't count as error)
                    skipped += 1
                    continue
            
            vehicle_data = VehicleCreate(
                brand=brand,
                model=safe_str(row.get('model', '')) or 'Unknown Model',
                chassis_number=chassis_number or 'Unknown Chassis',
                engine_number=engine_number or 'Unknown Engine',
                color=safe_str(row.get('color', '')) or 'Unknown Color',
                vehicle_number=vehicle_number or None,
                key_number=key_number or 'Unknown Key',
                inbound_location=safe_str(row.get('inbound_location', '')) or 'Unknown Location',
                page_number=safe_str(row.get('page_number', '')) or None
            )
            
            # Create vehicle with proper status
            vehicle_dict = vehicle_data.dict()
            vehicle_dict['status'] = status
            
            # Handle date_received field
            date_received_str = safe_str(row.get('date_received', ''))
            if date_received_str:
                try:
                    # Try to parse the date in various formats
                    date_received = parse_date_flexible(date_received_str)
                    vehicle_dict['date_received'] = date_received
                except Exception as e:
                    # If date parsing fails, use start of year (avoids inflating this-month stats)
                    vehicle_dict['date_received'] = datetime.now(timezone.utc)
            else:
                vehicle_dict['date_received'] = datetime.now(timezone.utc)
            
            vehicle = Vehicle(**vehicle_dict)
            
            # CROSS-REFERENCE: Check if customer mobile is provided
            customer_mobile = safe_str(row.get('customer_mobile', ''))
            customer_name = safe_str(row.get('customer_name', ''))
            customer_id = None
            
            if customer_mobile:
                # Find or create customer
                customer_id = await find_or_create_customer(
                    customer_mobile, 
                    {'name': customer_name or 'Unknown Customer'}, 
                    import_stats
                )
                vehicle.customer_id = customer_id
            
            await db.vehicles.insert_one(vehicle.dict())
            
            # CROSS-REFERENCE: Create sales record if sale data is provided
            sale_amount = safe_str(row.get('sale_amount', ''))
            payment_method = safe_str(row.get('payment_method', ''))
            
            if sale_amount and customer_id:
                try:
                    sale_record = Sale(
                        customer_id=customer_id,
                        vehicle_id=vehicle.id,
                        amount=float(sale_amount),
                        payment_method=payment_method.upper() or 'CASH',
                        sale_date=datetime.now(timezone.utc),
                        invoice_number=f"IMP-VEH-{vehicle.id[:8]}",
                        vehicle_brand=brand,
                        vehicle_model=vehicle_data.model,
                        vehicle_color=vehicle_data.color,
                        vehicle_chassis=chassis_number,
                        vehicle_engine=engine_number,
                        vehicle_registration=vehicle_number,
                        source="import",
                        created_by=user_id
                    )
                    await db.sales.insert_one(sale_record.dict())
                    import_stats['sales_created'] += 1
                    
                    # Update vehicle status to sold if sale is created
                    await db.vehicles.update_one(
                        {"id": vehicle.id},
                        {"$set": {"status": "sold"}}
                    )
                except Exception as sale_error:
                    print(f"Warning: Could not create sale record for vehicle {vehicle.id}: {sale_error}")
            
            # Track incomplete records (vehicles without customer linkage)
            if not customer_id and customer_mobile:
                incomplete_records.append({
                    "record_id": vehicle.id,
                    "row": idx + 2,
                    "missing_fields": ["customer_details"],
                    "data": row
                })
            
            successful += 1
            
        except Exception as e:
            failed += 1
            errors.append({
                "row": idx + 2,
                "data": row,
                "error": str(e)
            })
    
    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors
    import_job.cross_reference_stats = import_stats
    import_job.incomplete_records = incomplete_records
    
    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Import completed: {successful} successful, {failed} failed, {skipped} skipped (duplicates). Cross-referenced: {import_stats['customers_linked']} customers linked, {import_stats['customers_created']} customers created, {import_stats['sales_created']} sales created.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )

async def import_spare_parts_data(data: List[Dict], import_job: ImportJob, user_id: str) -> ImportResult:
    """Import spare parts with smart merge logic:
    - New part (part_number not found): insert everything
    - Existing part (part_number found):
        * All fields identical → skip
        * Missing/blank fields → fill only those
    Also: 28% GST is replaced with 18% on import and fixed in DB for existing records.
    """
    successful = 0
    failed = 0
    skipped = 0
    updated = 0
    errors = []
    gst_fixed = 0

    def fix_gst(val: float) -> float:
        """Replace 28% GST with 18%."""
        return 18.0 if val == 28.0 else val

    # ── Fix existing 28% GST records in DB ──────────────────────────────────
    fix_result = await db.spare_parts.update_many(
        {"gst_percentage": 28.0},
        {"$set": {"gst_percentage": 18.0}}
    )
    gst_fixed = fix_result.modified_count

    # Pre-fetch all existing parts by part_number
    existing_parts = {}
    async for p in db.spare_parts.find({}, {"_id": 0}):
        if p.get("part_number"):
            existing_parts[p["part_number"].strip()] = p

    for idx, row in enumerate(data):
        try:
            required_fields = ['name', 'part_number', 'brand', 'quantity', 'unit_price']
            for field in required_fields:
                if not row.get(field):
                    raise ValueError(f"{field} is required")

            part_number = str(row['part_number']).strip()
            name        = str(row['name']).strip()
            brand       = str(row['brand']).strip()
            quantity    = int(row['quantity'])
            unit_price  = float(row['unit_price'])
            unit        = str(row.get('unit', 'Nos')).strip()
            hsn_sac     = str(row.get('hsn_sac', '') or '').strip() or None
            gst_pct     = fix_gst(float(row.get('gst_percentage', 18.0) or 18.0))
            compat      = str(row.get('compatible_models', '') or '').strip() or None
            threshold   = int(row.get('low_stock_threshold', 5) or 5)
            supplier    = str(row.get('supplier', '') or '').strip() or None

            # ── EXISTING PART ──────────────────────────────────────────────────
            if part_number in existing_parts:
                existing = existing_parts[part_number]

                incoming = {
                    'name':              name,
                    'brand':             brand,
                    'unit':              unit,
                    'unit_price':        unit_price,
                    'hsn_sac':           hsn_sac,
                    'gst_percentage':    gst_pct,
                    'compatible_models': compat,
                    'supplier':          supplier,
                }

                update_fields = {}
                for field, val in incoming.items():
                    if val is not None and val != '' and not existing.get(field):
                        update_fields[field] = val

                # Quantity: add incoming to existing (restocking logic)
                if quantity and quantity > 0 and existing.get('quantity', 0) == 0:
                    update_fields['quantity'] = quantity

                if not update_fields:
                    skipped += 1
                    continue

                await db.spare_parts.update_one({"part_number": part_number}, {"$set": update_fields})
                existing_parts[part_number] = {**existing, **update_fields}
                updated += 1
                successful += 1
                continue

            # ── NEW PART ───────────────────────────────────────────────────────
            spare_part_data = SparePartCreate(
                name=name, part_number=part_number, brand=brand,
                quantity=quantity, unit=unit, unit_price=unit_price,
                hsn_sac=hsn_sac, gst_percentage=gst_pct,
                compatible_models=compat, low_stock_threshold=threshold, supplier=supplier
            )
            spare_part = SparePart(**spare_part_data.dict())
            await db.spare_parts.insert_one(spare_part.dict())
            existing_parts[part_number] = spare_part.dict()
            successful += 1

        except Exception as e:
            failed += 1
            errors.append({"row": idx + 2, "data": row, "error": str(e)})

    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Spare parts import: {successful} processed ({successful - updated} new, {updated} updated with missing fields), {failed} failed, {skipped} skipped (no new data). {gst_fixed} existing records fixed (28% → 18% GST).",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors
    )

async def import_services_data(data: List[Dict], import_job: ImportJob, user_id: str) -> ImportResult:
    """Import services (job cards) with smart merge logic:
    - Duplicate check: same customer_id + vehicle_number + service_type + amount → skip
    - If existing service found with same key but missing fields → fill only missing ones
    Also: fix any 28% GST items on existing service bills → 18%.
    """
    successful = 0
    failed = 0
    skipped = 0
    updated = 0
    errors = []
    incomplete_records = []
    import_stats = {'customers_linked': 0, 'customers_created': 0, 'vehicles_linked': 0, 'services_updated': 0}

    # ── Fix 28% GST on existing service bill items ───────────────────────────
    gst_fixed_bills = 0
    async for bill in db.service_bills.find({"items.gst_percentage": 28.0}, {"_id": 0, "id": 1, "items": 1}):
        fixed_items = []
        changed = False
        for item in bill.get("items", []):
            if item.get("gst_percentage") == 28.0:
                item = dict(item)
                item["gst_percentage"] = 18.0
                # Recalculate tax
                rate = item.get("rate", 0)
                qty  = item.get("quantity", 1)
                subtotal = rate * qty
                cgst = round(subtotal * 0.09, 2)
                sgst = round(subtotal * 0.09, 2)
                item["cgst"] = cgst
                item["sgst"] = sgst
                item["total"] = round(subtotal + cgst + sgst, 2)
                changed = True
            fixed_items.append(item)
        if changed:
            await db.service_bills.update_one({"id": bill["id"]}, {"$set": {"items": fixed_items}})
            gst_fixed_bills += 1

    for idx, row in enumerate(data):
        try:
            customer_mobile = (row.get('customer_mobile') or '').strip()
            customer_name   = (row.get('customer_name') or '').strip()
            vehicle_number  = (row.get('vehicle_number') or '').strip()
            chassis_number  = (row.get('chassis_number') or '').strip()
            service_type    = (row.get('service_type') or 'general_service').strip()
            description     = (row.get('description') or '').strip() or 'Imported service'
            amount          = float(row.get('amount', 0) or 0)

            if not customer_mobile and not vehicle_number and not chassis_number:
                raise ValueError("Either customer_mobile or vehicle identifiers required")

            # Resolve customer
            customer_id = None
            if customer_mobile:
                customer_id = await find_or_create_customer(
                    customer_mobile, {'name': customer_name or 'Unknown Customer'}, import_stats
                )

            # Resolve vehicle
            vehicle_id = None
            vehicle_brand = vehicle_model = vehicle_year = None
            if vehicle_number or chassis_number:
                vehicle = await find_vehicle_by_identifiers(vehicle_number, chassis_number)
                if vehicle:
                    vehicle_id    = vehicle.get('id')
                    vehicle_number = vehicle.get('vehicle_number', vehicle_number)
                    vehicle_brand  = vehicle.get('brand')
                    vehicle_model  = vehicle.get('model')
                    vehicle_year   = vehicle.get('year')
                    import_stats['vehicles_linked'] += 1
                    if not customer_id and vehicle.get('customer_id'):
                        customer_id = vehicle.get('customer_id')
                        import_stats['customers_linked'] += 1

            # Fallback vehicle details from CSV
            vehicle_brand = vehicle_brand or (row.get('vehicle_brand') or '').strip() or None
            vehicle_model = vehicle_model or (row.get('vehicle_model') or '').strip() or None
            vehicle_year  = vehicle_year  or (row.get('vehicle_year')  or '').strip() or None

            if not customer_id:
                customer_id = await find_or_create_customer(
                    f"AUTO-{str(uuid.uuid4())[:8]}",
                    {'name': customer_name or 'Unknown Customer'},
                    import_stats
                )
                incomplete_records.append({"row": idx + 2, "missing_fields": ["customer_mobile"], "data": row})

            # ── DEDUP CHECK ────────────────────────────────────────────────────
            dup_key = {
                "customer_id":  customer_id,
                "vehicle_number": vehicle_number or chassis_number or 'Unknown',
                "service_type": service_type,
                "amount":       amount
            }
            existing_service = await db.services.find_one(dup_key)

            if existing_service:
                # Fill only missing fields on existing record
                update_fields = {}
                if description and not existing_service.get('description'):
                    update_fields['description'] = description
                if vehicle_brand and not existing_service.get('vehicle_brand'):
                    update_fields['vehicle_brand'] = vehicle_brand
                if vehicle_model and not existing_service.get('vehicle_model'):
                    update_fields['vehicle_model'] = vehicle_model
                if vehicle_year and not existing_service.get('vehicle_year'):
                    update_fields['vehicle_year'] = vehicle_year
                if vehicle_id and not existing_service.get('vehicle_id'):
                    update_fields['vehicle_id'] = vehicle_id

                if not update_fields:
                    skipped += 1
                    continue

                await db.services.update_one({"id": existing_service['id']}, {"$set": update_fields})
                import_stats['services_updated'] += 1
                updated += 1
                successful += 1
                continue

            # ── NEW SERVICE ────────────────────────────────────────────────────
            service_data = ServiceCreate(
                customer_id=customer_id,
                vehicle_id=vehicle_id,
                vehicle_number=vehicle_number or chassis_number or 'Unknown',
                service_type=service_type,
                description=description,
                amount=amount
            )

            count = await db.services.count_documents({})
            job_card_number = f"JOB-{count + 1:06d}"

            service_dict = service_data.dict()
            service_dict['job_card_number'] = job_card_number
            service_dict['created_by'] = user_id

            reg_date = (row.get('registration_date') or '').strip()
            if reg_date:
                try:
                    from dateutil import parser as date_parser
                    service_dict['service_date'] = date_parser.parse(reg_date)
                except Exception:
                    pass  # leave service_date unset — don't default to today

            if vehicle_brand: service_dict['vehicle_brand'] = vehicle_brand
            if vehicle_model: service_dict['vehicle_model'] = vehicle_model
            if vehicle_year:  service_dict['vehicle_year']  = vehicle_year

            service = Service(**service_dict)
            await db.services.insert_one(service.dict())
            successful += 1

        except Exception as e:
            failed += 1
            errors.append({"row": idx + 2, "data": row, "error": str(e)})

    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors
    import_job.cross_reference_stats = import_stats
    import_job.incomplete_records = incomplete_records

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Services import: {successful} processed ({successful - updated} new, {updated} updated with missing fields), {failed} failed, {skipped} skipped (identical). Customers: {import_stats['customers_linked']} linked, {import_stats['customers_created']} created. Vehicles: {import_stats['vehicles_linked']} linked. {gst_fixed_bills} service bill(s) fixed (28% → 18% GST).",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )


    successful = 0
    failed = 0
    skipped = 0
    errors = []
    incomplete_records = []
    import_stats = {
        'customers_linked': 0,
        'customers_created': 0,
        'vehicles_linked': 0
    }
    
    for idx, row in enumerate(data):
        try:
            # Validate required fields (relaxed - only need mobile or vehicle identifier)
            customer_mobile = (row.get('customer_mobile') or '').strip()
            customer_name = (row.get('customer_name') or '').strip()
            vehicle_number = (row.get('vehicle_number') or '').strip()
            chassis_number = (row.get('chassis_number') or '').strip()
            service_type = (row.get('service_type') or 'general_service').strip()
            amount = float(row.get('amount', 0) or 0)
            
            if not customer_mobile and not vehicle_number and not chassis_number:
                raise ValueError("Either customer_mobile or vehicle identifiers (vehicle_number/chassis_number) must be provided")
            
            # CROSS-REFERENCE: Find or create customer
            customer_id = None
            if customer_mobile:
                customer_id = await find_or_create_customer(
                    customer_mobile,
                    {'name': customer_name or 'Unknown Customer'},
                    import_stats
                )
            
            # CROSS-REFERENCE: Find vehicle by identifiers
            vehicle_id = None
            vehicle_brand = None
            vehicle_model = None
            vehicle_year = None
            
            if vehicle_number or chassis_number:
                vehicle = await find_vehicle_by_identifiers(vehicle_number, chassis_number)
                if vehicle:
                    vehicle_id = vehicle.get('id')
                    vehicle_number = vehicle.get('vehicle_number', vehicle_number)
                    vehicle_brand = vehicle.get('brand')
                    vehicle_model = vehicle.get('model')
                    vehicle_year = vehicle.get('year')
                    import_stats['vehicles_linked'] += 1
                    
                    # If no customer was found by mobile, try to get from vehicle
                    if not customer_id and vehicle.get('customer_id'):
                        customer_id = vehicle.get('customer_id')
                        import_stats['customers_linked'] += 1
            
            # If vehicle not found in database, use CSV-provided vehicle details
            if not vehicle_brand:
                vehicle_brand = (row.get('vehicle_brand') or '').strip() or None
            if not vehicle_model:
                vehicle_model = (row.get('vehicle_model') or '').strip() or None
            if not vehicle_year:
                vehicle_year = (row.get('vehicle_year') or '').strip() or None
            
            # If still no customer, create a placeholder
            if not customer_id:
                customer_id = await find_or_create_customer(
                    f"AUTO-{str(uuid.uuid4())[:8]}",
                    {'name': customer_name or 'Unknown Customer'},
                    import_stats
                )
                incomplete_records.append({
                    "row": idx + 2,
                    "missing_fields": ["customer_mobile"],
                    "data": row
                })
            
            # Check for duplicate service (same customer, vehicle, service_type, and similar amount)
            duplicate_check = {
                "customer_id": customer_id,
                "vehicle_number": vehicle_number or chassis_number or 'Unknown',
                "service_type": service_type,
                "amount": amount
            }
            existing_service = await db.services.find_one(duplicate_check)
            if existing_service:
                # Skip duplicate service
                skipped += 1
                continue
            
            service_data = ServiceCreate(
                customer_id=customer_id,
                vehicle_id=vehicle_id,
                vehicle_number=vehicle_number or chassis_number or 'Unknown',
                service_type=service_type,
                description=(row.get('description') or '').strip() or 'Imported service',
                amount=amount
            )
            
            # Generate job card number
            count = await db.services.count_documents({})
            job_card_number = f"JOB-{count + 1:06d}"
            
            service_dict = service_data.dict()
            service_dict['job_card_number'] = job_card_number
            service_dict['created_by'] = user_id
            
            # Handle registration_date (service_date)
            registration_date = (row.get('registration_date') or '').strip()
            if registration_date:
                try:
                    from dateutil import parser as date_parser
                    parsed_date = date_parser.parse(registration_date)
                    service_dict['service_date'] = parsed_date
                except Exception:
                    # If parsing fails, use current datetime
                    service_dict['service_date'] = datetime.now(timezone.utc)
            
            # Add vehicle details for imported services (so they can be displayed even if vehicle is deleted)
            if vehicle_brand:
                service_dict['vehicle_brand'] = vehicle_brand
            if vehicle_model:
                service_dict['vehicle_model'] = vehicle_model
            if vehicle_year:
                service_dict['vehicle_year'] = vehicle_year
            
            service = Service(**service_dict)
            
            await db.services.insert_one(service.dict())
            successful += 1
            
        except Exception as e:
            failed += 1
            errors.append({
                "row": idx + 2,
                "data": row,
                "error": str(e)
            })
    
    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors
    import_job.cross_reference_stats = import_stats
    import_job.incomplete_records = incomplete_records
    
    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Import completed: {successful} successful, {failed} failed, {skipped} skipped (duplicates). Cross-referenced: {import_stats['customers_linked']} customers linked, {import_stats['customers_created']} customers created, {import_stats['vehicles_linked']} vehicles linked.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )

async def import_service_bills_data(data: List[Dict], import_job: ImportJob, user_id: str) -> ImportResult:
    """Import service bills from CSV/Excel.
    Each row = one bill with a single line item. Multiple rows with the same
    job_card_number (or customer+date combo) are grouped into one bill.
    """
    successful = 0
    failed = 0
    skipped = 0
    errors = []
    incomplete_records = []
    import_stats = {'customers_linked': 0, 'customers_created': 0}

    # Group rows by bill key so multiple items land on one bill
    bill_groups: Dict[str, List[Dict]] = {}
    for idx, row in enumerate(data):
        jcn = (row.get('job_card_number') or '').strip()
        mobile = (row.get('customer_mobile') or '').strip()
        date = (row.get('bill_date') or '').strip()
        # Key: prefer job card number, fall back to mobile+date
        key = jcn if jcn else f"{mobile}__{date}__{idx}"
        bill_groups.setdefault(key, []).append((idx, row))

    for key, rows in bill_groups.items():
        idx0, first = rows[0]
        try:
            customer_name = (first.get('customer_name') or '').strip()
            customer_mobile = (first.get('customer_mobile') or '').strip()
            vehicle_number = (first.get('vehicle_number') or '').strip()
            vehicle_brand = (first.get('vehicle_brand') or '').strip() or None
            vehicle_model = (first.get('vehicle_model') or '').strip() or None
            job_card_number = (first.get('job_card_number') or '').strip() or None
            status = (first.get('status') or 'pending').strip().lower()
            if status not in ('paid', 'pending', 'cancelled'):
                status = 'pending'

            # Resolve customer
            customer_id = None
            if customer_mobile:
                existing_cust = await db.customers.find_one({"mobile": customer_mobile})
                if existing_cust:
                    customer_id = existing_cust.get('id')
                    import_stats['customers_linked'] += 1
                else:
                    # Create minimal customer
                    from models import Customer  # noqa – already imported via *
                    new_cust_id = str(uuid.uuid4())
                    await db.customers.insert_one({
                        "id": new_cust_id, "name": customer_name or "Unknown",
                        "mobile": customer_mobile, "created_at": datetime.now(timezone.utc)
                    })
                    customer_id = new_cust_id
                    import_stats['customers_created'] += 1

            # Check duplicate bill by job card number
            if job_card_number:
                existing_bill = await db.service_bills.find_one({"job_card_number": job_card_number})
                if existing_bill:
                    skipped += 1
                    continue

            # Build bill items from all rows in this group
            items = []
            subtotal = 0.0
            total_cgst = 0.0
            total_sgst = 0.0

            for _, row in rows:
                desc = (row.get('item_description') or '').strip() or 'Service'
                hsn = (row.get('item_hsn') or '').strip()
                try:
                    qty = float(row.get('item_qty') or 1)
                    rate = float(row.get('item_rate') or 0)
                    gst_pct = float(row.get('item_gst_percent') or 18)
                except (ValueError, TypeError):
                    qty, rate, gst_pct = 1, 0, 18

                item_subtotal = qty * rate
                cgst = round(item_subtotal * gst_pct / 200, 2)
                sgst = round(item_subtotal * gst_pct / 200, 2)
                item_total = round(item_subtotal + cgst + sgst, 2)

                items.append({
                    "description": desc, "hsn_sac": hsn,
                    "quantity": qty, "rate": rate,
                    "gst_percentage": gst_pct,
                    "cgst": cgst, "sgst": sgst,
                    "total": item_total
                })
                subtotal += item_subtotal
                total_cgst += cgst
                total_sgst += sgst

            # Use total_amount from CSV if provided and items are empty/zero
            csv_total = None
            try:
                csv_total = float(first.get('total_amount') or 0) or None
            except (ValueError, TypeError):
                pass

            total_tax = round(total_cgst + total_sgst, 2)
            total_amount = csv_total if (csv_total and not items) else round(subtotal + total_tax, 2)

            # If no items built, create a single line from total_amount
            if not items and total_amount:
                items = [{"description": "Service (imported)", "hsn_sac": "", "quantity": 1,
                          "rate": total_amount, "gst_percentage": 0, "cgst": 0, "sgst": 0, "total": total_amount}]
                subtotal = total_amount
                total_tax = 0

            # Bill date
            bill_date = datetime.now(timezone.utc)
            raw_date = (first.get('bill_date') or '').strip()
            if raw_date:
                try:
                    from dateutil import parser as date_parser
                    bill_date = date_parser.parse(raw_date)
                except Exception:
                    pass

            # Generate bill number
            seq = await db.service_bills.count_documents({})
            bill_number = f"SB-{seq + 1:06d}"

            bill = ServiceBill(
                bill_number=bill_number,
                job_card_number=job_card_number,
                customer_id=customer_id,
                customer_name=customer_name or None,
                customer_mobile=customer_mobile or None,
                vehicle_number=vehicle_number or None,
                vehicle_brand=vehicle_brand,
                vehicle_model=vehicle_model,
                items=items,
                subtotal=round(subtotal, 2),
                total_discount=0,
                total_cgst=round(total_cgst, 2),
                total_sgst=round(total_sgst, 2),
                total_tax=total_tax,
                total_amount=total_amount,
                bill_date=bill_date,
                status=status,
                created_by=user_id
            )
            await db.service_bills.insert_one(bill.dict())
            successful += 1

        except Exception as e:
            failed += 1
            errors.append({"row": idx0 + 2, "error": str(e)})

    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors
    import_job.cross_reference_stats = import_stats
    import_job.incomplete_records = incomplete_records

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Service bills import: {successful} bills created, {failed} failed, {skipped} skipped (duplicate job card). {import_stats['customers_linked']} customers linked, {import_stats['customers_created']} customers created.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Service bills import: {successful} bills created, {failed} failed, {skipped} skipped (duplicate job card). {import_stats['customers_linked']} customers linked, {import_stats['customers_created']} customers created.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )

async def import_registrations_data(data: List[Dict], import_job: ImportJob, user_id: str) -> ImportResult:
    """Import service registrations (REG-XXXXXX). Auto-links or creates customers."""
    successful = 0
    failed = 0
    skipped = 0
    errors = []
    incomplete_records = []
    import_stats = {'customers_linked': 0, 'customers_created': 0}

    for idx, row in enumerate(data):
        try:
            customer_name = (row.get('customer_name') or '').strip()
            customer_mobile = (row.get('customer_mobile') or '').strip()
            vehicle_number = (row.get('vehicle_number') or '').strip()

            if not customer_mobile:
                raise ValueError("customer_mobile is required")
            if not vehicle_number:
                raise ValueError("vehicle_number is required")

            # Dedup: skip if same mobile + vehicle_number already registered
            existing_reg = await db.registrations.find_one({
                "customer_mobile": customer_mobile,
                "vehicle_number": vehicle_number
            })
            if existing_reg:
                skipped += 1
                continue

            # Find or create customer
            existing_customer = await db.customers.find_one({"mobile": customer_mobile})
            if existing_customer:
                customer_id = existing_customer["id"]
                updates = {}
                if customer_name and existing_customer.get("name") != customer_name:
                    updates["name"] = customer_name
                existing_type = existing_customer.get("customer_type", "sales")
                if existing_type == "sales":
                    updates["customer_type"] = "both"
                elif not existing_type:
                    updates["customer_type"] = "service"
                if updates:
                    await db.customers.update_one({"id": customer_id}, {"$set": updates})
                import_stats['customers_linked'] += 1
            else:
                customer_id = str(uuid.uuid4())
                await db.customers.insert_one({
                    "id": customer_id,
                    "name": customer_name or "Unknown",
                    "mobile": customer_mobile,
                    "address": (row.get('customer_address') or '').strip(),
                    "customer_type": "service",
                    "created_at": datetime.now(timezone.utc)
                })
                import_stats['customers_created'] += 1

            # Parse registration_date — store None if not provided, don't default to today
            reg_date = None
            raw_date = (row.get('registration_date') or '').strip()
            if raw_date:
                try:
                    from dateutil import parser as date_parser
                    reg_date = date_parser.parse(raw_date)
                except Exception:
                    incomplete_records.append({"row": idx + 2, "missing_fields": ["registration_date (invalid format)"]})

            seq = await next_sequence("registrations")
            registration_number = f"REG-{seq:06d}"

            reg_dict = {
                "id": str(uuid.uuid4()),
                "registration_number": registration_number,
                "customer_id": customer_id,
                "customer_name": customer_name or "Unknown",
                "customer_mobile": customer_mobile,
                "customer_address": (row.get('customer_address') or '').strip() or None,
                "vehicle_number": vehicle_number,
                "vehicle_brand": (row.get('vehicle_brand') or '').strip() or None,
                "vehicle_model": (row.get('vehicle_model') or '').strip() or None,
                "vehicle_year": (row.get('vehicle_year') or '').strip() or None,
                "chassis_number": (row.get('chassis_number') or '').strip() or None,
                "engine_number": (row.get('engine_number') or '').strip() or None,
                "registration_date": reg_date,  # None if not provided
                "created_by": user_id,
                "created_at": datetime.now(timezone.utc)
            }
            await db.registrations.insert_one(reg_dict)
            successful += 1

            missing = []
            if not customer_name:
                missing.append("customer_name")
            if not raw_date:
                missing.append("registration_date")
            if missing:
                incomplete_records.append({"row": idx + 2, "missing_fields": missing})

        except Exception as e:
            failed += 1
            errors.append({"row": idx + 2, "error": str(e)})

    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors
    import_job.cross_reference_stats = import_stats
    import_job.incomplete_records = incomplete_records

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Registrations import: {successful} created, {failed} failed, {skipped} skipped (duplicates by mobile+vehicle). {import_stats['customers_linked']} customers linked, {import_stats['customers_created']} customers created.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Registrations import: {successful} created, {failed} failed, {skipped} skipped (duplicates by mobile+vehicle). {import_stats['customers_linked']} customers linked, {import_stats['customers_created']} customers created.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )

async def import_service_customers_data(data: List[Dict], import_job: ImportJob, user_id: str) -> ImportResult:
    """Import service-only customers. Sets customer_type='service'.
    Skips if mobile already exists. Attaches vehicle_info if provided.
    """
    successful = 0
    failed = 0
    skipped = 0
    errors = []
    incomplete_records = []
    import_stats = {'vehicles_linked': 0}

    # Pre-fetch existing mobiles
    existing_mobiles = set()
    async for c in db.customers.find({}, {"mobile": 1}):
        if c.get("mobile"):
            existing_mobiles.add(str(c["mobile"]))

    existing_vehicles = {}
    async for v in db.vehicles.find({}, {"id": 1, "chassis_number": 1, "vehicle_number": 1}):
        if v.get("chassis_number"):
            existing_vehicles[v["chassis_number"]] = v
        if v.get("vehicle_number"):
            existing_vehicles[v["vehicle_number"]] = v

    for idx, row in enumerate(data):
        try:
            mobile = (safe_str(row.get('mobile', '')) or safe_str(row.get('phone', ''))).strip()
            name = safe_str(row.get('name', '')).strip() or "Unknown"

            if not mobile:
                raise ValueError("mobile is required")

            if mobile in existing_mobiles:
                # Update existing customer to include service type
                existing = await db.customers.find_one({"mobile": mobile})
                if existing:
                    current_type = existing.get("customer_type", "sales")
                    if current_type == "sales":
                        await db.customers.update_one(
                            {"id": existing["id"]},
                            {"$set": {"customer_type": "both"}}
                        )
                skipped += 1
                continue
            existing_mobiles.add(mobile)

            customer_id = str(uuid.uuid4())
            customer_dict = {
                "id": customer_id,
                "name": name,
                "mobile": mobile,
                "phone": safe_str(row.get('phone', '')).strip() or None,
                "email": safe_str(row.get('email', '')).strip() or None,
                "address": safe_str(row.get('address', '')).strip() or "",
                "customer_type": "service",
                "created_at": datetime.now(timezone.utc),
                "created_by": user_id
            }

            # Vehicle info
            chassis = safe_str(row.get('chassis_number', '')).strip()
            vehicle_no = safe_str(row.get('vehicle_number', '')).strip()
            vehicle_info = {}
            if any([row.get('vehicle_brand'), row.get('vehicle_model'), vehicle_no, chassis]):
                vehicle_info = {
                    "brand": safe_str(row.get('vehicle_brand', '')).strip(),
                    "model": safe_str(row.get('vehicle_model', '')).strip(),
                    "year": safe_str(row.get('vehicle_year', '')).strip(),
                    "vehicle_number": vehicle_no,
                    "chassis_number": chassis,
                    "engine_number": safe_str(row.get('engine_number', '')).strip(),
                }
                customer_dict["vehicle_info"] = vehicle_info

            # Link to existing vehicle in stock
            matched = existing_vehicles.get(chassis) or existing_vehicles.get(vehicle_no)
            if matched:
                await db.vehicles.update_one(
                    {"id": matched["id"]},
                    {"$set": {"customer_id": customer_id}}
                )
                import_stats["vehicles_linked"] += 1

            await db.customers.insert_one(customer_dict)
            successful += 1

            missing = []
            if not safe_str(row.get('address', '')).strip():
                missing.append("address")
            if not vehicle_info:
                missing.append("vehicle details")
            if missing:
                incomplete_records.append({"row": idx + 2, "missing_fields": missing})

        except Exception as e:
            failed += 1
            errors.append({"row": idx + 2, "error": str(e)})

    import_job.successful_records = successful
    import_job.failed_records = failed
    import_job.skipped_records = skipped
    import_job.processed_records = successful + failed + skipped
    import_job.errors = errors
    import_job.cross_reference_stats = import_stats
    import_job.incomplete_records = incomplete_records

    return ImportResult(
        job_id=import_job.id,
        status="completed",
        message=f"Service customers import: {successful} created, {failed} failed, {skipped} already existed (type updated to 'both'). {import_stats['vehicles_linked']} vehicles linked.",
        total_records=len(data),
        successful_records=successful,
        failed_records=failed,
        skipped_records=skipped,
        errors=errors,
        cross_reference_stats=import_stats,
        incomplete_records=incomplete_records
    )

# Duplicate Detection and Cleanup endpoints
@api_router.get("/duplicates/detect")
async def detect_duplicates(current_user: User = Depends(get_current_user)):
    """Detect duplicate records across all collections"""
    duplicates = {
        "vehicles": {},
        "customers": {},
        "summary": {}
    }
    
    # Find duplicate vehicles by chassis_number
    vehicle_pipeline = [
        {"$match": {"chassis_number": {"$ne": None, "$ne": ""}}},
        {"$group": {
            "_id": "$chassis_number",
            "count": {"$sum": 1},
            "ids": {"$push": "$id"},
            "records": {"$push": "$$ROOT"}
        }},
        {"$match": {"count": {"$gt": 1}}}
    ]
    
    vehicle_duplicates = await db.vehicles.aggregate(vehicle_pipeline).to_list(1000)
    
    # Find duplicate customers by mobile
    customer_pipeline = [
        {"$match": {"mobile": {"$ne": None, "$ne": ""}}},
        {"$group": {
            "_id": "$mobile", 
            "count": {"$sum": 1},
            "ids": {"$push": "$id"},
            "records": {"$push": "$$ROOT"}
        }},
        {"$match": {"count": {"$gt": 1}}}
    ]
    
    customer_duplicates = await db.customers.aggregate(customer_pipeline).to_list(1000)
    
    # Process vehicle duplicates
    total_vehicle_duplicates = 0
    for duplicate in vehicle_duplicates:
        chassis_no = duplicate["_id"]
        duplicates["vehicles"][chassis_no] = {
            "count": duplicate["count"],
            "ids": duplicate["ids"],
            "records": [{"id": r["id"], "brand": r.get("brand"), "model": r.get("model"), "color": r.get("color")} for r in duplicate["records"]]
        }
        total_vehicle_duplicates += duplicate["count"] - 1  # Subtract 1 to keep original
    
    # Process customer duplicates
    total_customer_duplicates = 0
    for duplicate in customer_duplicates:
        mobile = duplicate["_id"]
        duplicates["customers"][mobile] = {
            "count": duplicate["count"],
            "ids": duplicate["ids"],
            "records": [{"id": r["id"], "name": r.get("name"), "email": r.get("email")} for r in duplicate["records"]]
        }
        total_customer_duplicates += duplicate["count"] - 1  # Subtract 1 to keep original
    
    duplicates["summary"] = {
        "total_vehicle_duplicates": total_vehicle_duplicates,
        "total_customer_duplicates": total_customer_duplicates,
        "vehicle_chassis_groups": len(vehicle_duplicates),
        "customer_mobile_groups": len(customer_duplicates)
    }
    
    return duplicates

@api_router.post("/duplicates/cleanup")
async def cleanup_duplicates(current_user: User = Depends(get_current_user)):
    """Remove duplicate records, keeping the oldest one in each group"""
    
    cleanup_results = {
        "vehicles_removed": 0,
        "customers_removed": 0,
        "removed_vehicle_ids": [],
        "removed_customer_ids": []
    }
    
    # Clean up duplicate vehicles by chassis_number
    vehicle_pipeline = [
        {"$match": {"chassis_number": {"$ne": None, "$ne": ""}}},
        {"$group": {
            "_id": "$chassis_number",
            "count": {"$sum": 1},
            "records": {"$push": "$$ROOT"}
        }},
        {"$match": {"count": {"$gt": 1}}}
    ]
    
    vehicle_duplicates = await db.vehicles.aggregate(vehicle_pipeline).to_list(1000)
    
    for duplicate_group in vehicle_duplicates:
        records = duplicate_group["records"]
        # Sort by created_at to keep the oldest
        records.sort(key=lambda x: x.get("date_received", datetime.now(timezone.utc)))
        
        # Keep the first (oldest) and remove the rest
        to_remove = records[1:]
        for record in to_remove:
            # Check if vehicle is not associated with sales/services
            sales_count = await db.sales.count_documents({"vehicle_id": record["id"]})
            services_count = await db.services.count_documents({"vehicle_id": record["id"]})
            
            if sales_count == 0 and services_count == 0:
                await db.vehicles.delete_one({"id": record["id"]})
                cleanup_results["vehicles_removed"] += 1
                cleanup_results["removed_vehicle_ids"].append(record["id"])
    
    # Clean up duplicate customers by mobile
    customer_pipeline = [
        {"$match": {"mobile": {"$ne": None, "$ne": ""}}},
        {"$group": {
            "_id": "$mobile",
            "count": {"$sum": 1},
            "records": {"$push": "$$ROOT"}
        }},
        {"$match": {"count": {"$gt": 1}}}
    ]
    
    customer_duplicates = await db.customers.aggregate(customer_pipeline).to_list(1000)
    
    for duplicate_group in customer_duplicates:
        records = duplicate_group["records"]
        # Sort by created_at to keep the oldest
        records.sort(key=lambda x: x.get("created_at", datetime.now(timezone.utc)))
        
        # Keep the first (oldest) and remove the rest
        to_remove = records[1:]
        for record in to_remove:
            # Check if customer is not associated with sales/services
            sales_count = await db.sales.count_documents({"customer_id": record["id"]})
            services_count = await db.services.count_documents({"customer_id": record["id"]})
            
            if sales_count == 0 and services_count == 0:
                await db.customers.delete_one({"id": record["id"]})
                cleanup_results["customers_removed"] += 1
                cleanup_results["removed_customer_ids"].append(record["id"])
    
    return cleanup_results

# Include the router in the main app
# Backup Service Class
class BackupService:
    def __init__(self, db, backup_config: BackupConfig):
        self.db = db
        self.config = backup_config
        self.backup_root = Path(backup_config.backup_location)
        self.backup_root.mkdir(parents=True, exist_ok=True)
    
    async def create_backup(self, user_id: str, backup_type: str = "manual", export_format: str = "json") -> BackupJob:
        """Create a new backup job with specified format"""
        job = BackupJob(
            status="running",
            start_time=datetime.utcnow(),
            created_by=user_id,
            backup_type=backup_type
        )
        
        try:
            # Create backup directory
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_dir = self.backup_root / f"backup_{timestamp}"
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            # Backup all collections
            collections = ['users', 'customers', 'vehicles', 'sales', 'services', 'spare_parts', 'spare_part_bills']
            total_records = 0
            records_by_collection = {}
            collection_data = {}
            
            for collection_name in collections:
                collection = getattr(self.db, collection_name)
                documents = await collection.find().to_list(length=None)
                
                # Convert ObjectId to string for JSON serialization
                for doc in documents:
                    if '_id' in doc:
                        doc['_id'] = str(doc['_id'])
                
                collection_data[collection_name] = documents
                record_count = len(documents)
                records_by_collection[collection_name] = record_count
                total_records += record_count
            
            # Create files based on export format
            if export_format.lower() == "excel":
                await self.create_excel_backup(backup_dir, collection_data, records_by_collection, user_id, backup_type)
            else:
                # Default JSON/CSV format
                await self.create_json_csv_backup(backup_dir, collection_data, records_by_collection, user_id, backup_type)
            
            # Compress backup if enabled
            final_path = str(backup_dir)
            backup_size = 0
            
            if self.config.compress_backups:
                if export_format.lower() == "excel":
                    zip_path = f"{backup_dir}_excel.zip"
                else:
                    zip_path = f"{backup_dir}.zip"
                
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file_path in backup_dir.rglob('*'):
                        if file_path.is_file():
                            zipf.write(file_path, file_path.relative_to(backup_dir))
                
                # Remove original directory and get compressed size
                shutil.rmtree(backup_dir)
                backup_size = os.path.getsize(zip_path) / (1024 * 1024)  # MB
                final_path = zip_path
            else:
                # Calculate directory size
                backup_size = sum(f.stat().st_size for f in backup_dir.rglob('*') if f.is_file()) / (1024 * 1024)
            
            # Update job with completion details
            job.status = "completed"
            job.end_time = datetime.utcnow()
            job.total_records = total_records
            job.backup_size_mb = round(backup_size, 2)
            job.backup_file_path = final_path
            job.records_backed_up = records_by_collection
            
        except Exception as e:
            job.status = "failed"
            job.end_time = datetime.utcnow()
            job.error_message = str(e)
        
        # Save job to database
        await self.db.backup_jobs.insert_one(job.dict())
        
        return job
    
    async def create_excel_backup(self, backup_dir: Path, collection_data: dict, records_by_collection: dict, user_id: str, backup_type: str):
        """Create Excel backup with multiple sheets"""
        try:
            # Create comprehensive Excel file
            excel_file = backup_dir / "backup_data.xlsx"
            
            # Create workbook and remove default sheet
            workbook = Workbook()
            workbook.remove(workbook.active)
            
            # Define header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Create sheets for each collection
            for collection_name, documents in collection_data.items():
                if not documents:
                    continue
                    
                # Create sheet
                sheet = workbook.create_sheet(title=collection_name.replace('_', ' ').title())
                
                try:
                    # Flatten nested data for Excel compatibility
                    flattened_documents = []
                    for doc in documents:
                        flat_doc = self.flatten_document(doc)
                        flattened_documents.append(flat_doc)
                    
                    # Convert to DataFrame for easier Excel manipulation
                    df = pd.DataFrame(flattened_documents)
                    
                    # Add data to sheet
                    for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                        for col_num, value in enumerate(row_data, 1):
                            # Convert non-serializable values to strings
                            if value is None:
                                value = ""
                            elif isinstance(value, (dict, list)):
                                value = str(value)
                            
                            cell = sheet.cell(row=row_num, column=col_num, value=str(value))
                            
                            # Apply header styling
                            if row_num == 1:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                    
                    # Auto-adjust column widths
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except (TypeError, AttributeError):
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width
                        
                except Exception as e:
                    logger.warning(f"Failed to create Excel sheet for {collection_name}: {e}")
                    # Create a simple sheet with error message
                    sheet.cell(row=1, column=1, value=f"Error processing {collection_name}")
                    sheet.cell(row=2, column=1, value=str(e))
            
            # Create summary sheet
            summary_sheet = workbook.create_sheet(title="Backup Summary", index=0)
            
            # Summary data with IST timezone
            current_utc = datetime.utcnow()
            ist_time = current_utc + timedelta(hours=5, minutes=30)
            
            summary_data = [
                ["Backup Information", ""],
                ["Backup Date", ist_time.strftime("%Y-%m-%d %H:%M:%S IST")],
                ["Backup Type", backup_type.title()],
                ["Created By", user_id],
                ["Total Records", sum(records_by_collection.values())],
                ["Export Format", "Excel Workbook"],
                ["Timezone", "IST (UTC+5:30)"],
                ["", ""],
                ["Collection Statistics", "Records"],
            ]
            
            # Add collection statistics
            for collection, count in records_by_collection.items():
                summary_data.append([collection.replace('_', ' ').title(), count])
            
            # Add summary data to sheet
            for row_num, (key, value) in enumerate(summary_data, 1):
                summary_sheet.cell(row=row_num, column=1, value=str(key))
                summary_sheet.cell(row=row_num, column=2, value=str(value))
                
                # Style headers
                if key in ["Backup Information", "Collection Statistics"]:
                    summary_sheet.cell(row=row_num, column=1).font = Font(bold=True, size=14)
                    summary_sheet.cell(row=row_num, column=2).font = Font(bold=True, size=14)
            
            # Auto-adjust summary sheet columns
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 15
            
            # Save workbook
            workbook.save(excel_file)
            logger.info(f"Excel backup created successfully: {excel_file}")
            
        except Exception as e:
            # Fallback to JSON if Excel creation fails
            logger.error(f"Excel backup creation failed: {e}, falling back to JSON")
            await self.create_json_csv_backup(backup_dir, collection_data, records_by_collection, user_id, backup_type)
    
    def flatten_document(self, doc: dict, prefix: str = "") -> dict:
        """Flatten nested dictionary for Excel compatibility"""
        flattened = {}
        
        for key, value in doc.items():
            new_key = f"{prefix}_{key}" if prefix else key
            
            if isinstance(value, dict):
                # Recursively flatten nested dictionaries
                nested_flattened = self.flatten_document(value, new_key)
                flattened.update(nested_flattened)
            elif isinstance(value, list):
                # Convert lists to comma-separated strings
                if value and isinstance(value[0], dict):
                    # For list of dictionaries, create a summary
                    flattened[new_key] = f"[{len(value)} items]"
                    # Add first item details if available
                    if value:
                        first_item = self.flatten_document(value[0], f"{new_key}_item1")
                        flattened.update(first_item)
                else:
                    # For simple lists, join as string
                    flattened[new_key] = ", ".join(str(item) for item in value)
            else:
                # Simple values
                flattened[new_key] = value
        
        return flattened
    
    async def create_json_csv_backup(self, backup_dir: Path, collection_data: dict, records_by_collection: dict, user_id: str, backup_type: str):
        """Create JSON and CSV backup files"""
        for collection_name, documents in collection_data.items():
            # Save as JSON
            json_file = backup_dir / f"{collection_name}.json"
            async with aiofiles.open(json_file, 'w') as f:
                await f.write(json.dumps(documents, default=str, indent=2))
            
            # Save as CSV if data exists
            if documents:
                try:
                    df = pd.DataFrame(documents)
                    csv_file = backup_dir / f"{collection_name}.csv"
                    df.to_csv(csv_file, index=False)
                except Exception as e:
                    logger.warning(f"Failed to create CSV for {collection_name}: {e}")
        
        # Create backup summary with IST timezone
        current_utc = datetime.utcnow()
        ist_time = current_utc + timedelta(hours=5, minutes=30)
        
        summary = {
            'backup_date': ist_time.isoformat(),
            'backup_date_utc': current_utc.isoformat(),
            'timezone': 'IST (UTC+5:30)',
            'total_records': sum(records_by_collection.values()),
            'records_by_collection': records_by_collection,
            'backup_type': backup_type,
            'created_by': user_id
        }
        
        summary_file = backup_dir / 'backup_summary.json'
        async with aiofiles.open(summary_file, 'w') as f:
            await f.write(json.dumps(summary, indent=2))
    
    async def get_backup_stats(self) -> BackupStats:
        """Get backup statistics"""
        jobs = await self.db.backup_jobs.find().to_list(length=None)
        
        total_backups = len(jobs)
        successful_backups = len([j for j in jobs if j['status'] == 'completed'])
        failed_backups = len([j for j in jobs if j['status'] == 'failed'])
        
        last_backup = None
        oldest_backup = None
        total_size = 0
        
        if jobs:
            sorted_jobs = sorted(jobs, key=lambda x: x['created_at'], reverse=True)
            last_backup = sorted_jobs[0]['created_at']
            oldest_backup = sorted_jobs[-1]['created_at']
            
            # Calculate total storage used
            for job in jobs:
                if job['status'] == 'completed':
                    total_size += job.get('backup_size_mb', 0)
        
        return BackupStats(
            total_backups=total_backups,
            successful_backups=successful_backups,
            failed_backups=failed_backups,
            last_backup_date=last_backup,
            oldest_backup_date=oldest_backup,
            total_storage_used_mb=round(total_size, 2)
        )
    
    async def cleanup_old_backups(self, retention_days: int):
        """Clean up backups older than retention period"""
        cutoff_date = datetime.utcnow() - timedelta(days=retention_days)
        
        # Remove old backup files
        for backup_path in self.backup_root.glob('backup_*'):
            if backup_path.stat().st_mtime < cutoff_date.timestamp():
                try:
                    if backup_path.is_file():
                        backup_path.unlink()
                    else:
                        shutil.rmtree(backup_path)
                    
                    # Update database to mark as cleaned up
                    await self.db.backup_jobs.update_one(
                        {"backup_file_path": str(backup_path)},
                        {"$set": {"cleaned_up": True}}
                    )
                except Exception as e:
                    logger.error(f"Failed to cleanup backup {backup_path}: {e}")

# Initialize backup service
backup_service = None

async def get_backup_service():
    """Get or create backup service instance"""
    global backup_service
    
    if backup_service is None:
        # Get backup config from database or create default
        config_doc = await db.backup_config.find_one()
        if not config_doc:
            # Create default config
            default_config = BackupConfig()
            await db.backup_config.insert_one(default_config.dict())
            config = default_config
        else:
            config = BackupConfig(**config_doc)
        
        backup_service = BackupService(db, config)
    
    return backup_service

# Backup API Endpoints
@app.get("/api/backup/config", response_model=BackupConfig)
async def get_backup_config(current_user: dict = Depends(verify_token)):
    """Get backup configuration"""
    config_doc = await db.backup_config.find_one()
    if not config_doc:
        # Create default config
        default_config = BackupConfig()
        await db.backup_config.insert_one(default_config.dict())
        return default_config
    return BackupConfig(**config_doc)

@app.put("/api/backup/config", response_model=BackupConfig)
async def update_backup_config(
    config_update: dict,
    current_user: dict = Depends(verify_token)
):
    """Update backup configuration"""
    config_update['updated_at'] = datetime.utcnow()
    
    result = await db.backup_config.update_one(
        {}, 
        {"$set": config_update},
        upsert=True
    )
    
    # Refresh backup service with new config
    global backup_service
    backup_service = None
    
    updated_config = await db.backup_config.find_one()
    return BackupConfig(**updated_config)

@app.post("/api/backup/create", response_model=BackupJob)
async def create_manual_backup(
    backup_create: BackupJobCreate,
    current_user: dict = Depends(verify_token)
):
    """Create a manual backup"""
    service = await get_backup_service()
    job = await service.create_backup(
        current_user['user_id'], 
        backup_create.backup_type,
        backup_create.export_format
    )
    return job

@app.get("/api/backup/jobs", response_model=List[BackupJob])
async def get_backup_jobs(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(verify_token)
):
    """Get backup job history"""
    jobs = await db.backup_jobs.find().skip(skip).limit(limit).sort("created_at", -1).to_list(length=None)
    return [BackupJob(**job) for job in jobs]

@app.get("/api/backup/stats", response_model=BackupStats)
async def get_backup_statistics(current_user: dict = Depends(verify_token)):
    """Get backup system statistics"""
    service = await get_backup_service()
    return await service.get_backup_stats()

@app.delete("/api/backup/cleanup")
async def cleanup_old_backups(
    retention_days: int = 30,
    current_user: dict = Depends(verify_token)
):
    """Clean up old backups"""
    service = await get_backup_service()
    await service.cleanup_old_backups(retention_days)
    return {"message": f"Cleanup completed for backups older than {retention_days} days"}

@app.get("/api/backup/download/{job_id}")
async def download_backup(
    job_id: str,
    current_user: dict = Depends(verify_token)
):
    """Download a backup file"""
    from fastapi.responses import FileResponse
    
    job = await db.backup_jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Backup job not found")
    
    if job['status'] != 'completed':
        raise HTTPException(status_code=400, detail="Backup is not completed")
    
    backup_path = Path(job['backup_file_path'])
    if not backup_path.exists():
        raise HTTPException(status_code=404, detail="Backup file not found")
    
    return FileResponse(
        path=backup_path,
        filename=backup_path.name,
        media_type='application/octet-stream'
    )

# ==================== Activity/Notifications System ====================

class ActivityType(str, Enum):
    SALE_CREATED = "sale_created"
    SERVICE_COMPLETED = "service_completed"
    SERVICE_CREATED = "service_created"
    VEHICLE_ADDED = "vehicle_added"
    VEHICLE_SOLD = "vehicle_sold"
    LOW_STOCK = "low_stock"
    CUSTOMER_ADDED = "customer_added"
    BACKUP_CREATED = "backup_created"

class Activity(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: ActivityType
    title: str
    description: str
    icon: str = "info"  # info, success, warning, error
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    read: bool = False
    metadata: Optional[Dict[str, Any]] = None

class ActivityCreate(BaseModel):
    type: ActivityType
    title: str
    description: str
    icon: str = "info"
    metadata: Optional[Dict[str, Any]] = None

async def create_activity(activity_data: ActivityCreate):
    """Helper function to create an activity"""
    activity = Activity(**activity_data.dict())
    await db.activities.insert_one(activity.dict())
    return activity

@app.get("/api/activities")
async def get_activities(
    limit: int = 20,
    skip: int = 0,
    unread_only: bool = False,
    current_user: dict = Depends(verify_token)
):
    """Get recent activities/notifications"""
    query = {}
    if unread_only:
        query["read"] = False
    
    activities = await db.activities.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    total = await db.activities.count_documents(query)
    unread_count = await db.activities.count_documents({"read": False})
    
    return {
        "activities": activities,
        "total": total,
        "unread_count": unread_count
    }

@app.post("/api/activities/{activity_id}/mark-read")
async def mark_activity_read(
    activity_id: str,
    current_user: dict = Depends(verify_token)
):
    """Mark an activity as read"""
    result = await db.activities.update_one(
        {"id": activity_id},
        {"$set": {"read": True}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Activity not found")
    
    return {"message": "Activity marked as read"}

@app.post("/api/activities/mark-all-read")
async def mark_all_activities_read(current_user: dict = Depends(verify_token)):
    """Mark all activities as read"""
    await db.activities.update_many(
        {"read": False},
        {"$set": {"read": True}}
    )
    
    return {"message": "All activities marked as read"}

@app.delete("/api/activities/{activity_id}")
async def delete_activity(
    activity_id: str,
    current_user: dict = Depends(verify_token)
):
    """Delete an activity"""
    result = await db.activities.delete_one({"id": activity_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Activity not found")
    
    return {"message": "Activity deleted"}

# ==================== End Activity/Notifications System ====================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "http://localhost:3000").split(","),
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
